"""Authenticated HTTP routes for the private local workbench."""

from __future__ import annotations

import base64
import binascii
import html
import json
import logging
import os
import secrets
import warnings
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Security, status
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from PIL import Image, UnidentifiedImageError
from pydantic import BaseModel, ConfigDict, Field, field_validator

from src.ace.workbench.engagement import (
    EngagementConflictError,
    EngagementDraft,
    EngagementG0Error,
    EngagementNotFoundServiceError,
    EngagementService,
    EngagementValidationError,
)
from src.ace.workbench.evidence_review import (
    EvidenceReviewConflictError,
    EvidenceReviewNotFoundError,
    EvidenceReviewService,
    EvidenceReviewValidationError,
)
from src.ace.workbench.relationship_review import (
    RelationshipReviewActor,
    RelationshipReviewActorKind,
    RelationshipReviewService,
)
from src.ace.workbench.storage import (
    CaptureAttemptConflictError,
    EvidenceReviewG0Error,
    NoReadyCurrentEngagementError,
    WorkbenchStore,
    resolve_data_dir,
    utc_now,
)


MAX_IMAGE_BYTES = 10 * 1024 * 1024
ALLOWED_MEDIA_TYPES = {"image/jpeg", "image/png", "image/webp", "image/gif"}
IMAGE_FORMATS = {
    "image/jpeg": "JPEG",
    "image/png": "PNG",
    "image/webp": "WEBP",
    "image/gif": "GIF",
}
basic_auth = HTTPBasic(auto_error=False)
router = APIRouter(prefix="/workbench")
logger = logging.getLogger(__name__)


class CaptureRequest(BaseModel):
    filename: str = Field(min_length=1, max_length=255)
    media_type: str
    data_base64: str = Field(min_length=1)
    capture_attempt_key: str | None = Field(
        default=None, min_length=8, max_length=128, pattern=r"^[A-Za-z0-9][A-Za-z0-9_.-]*$"
    )

    @field_validator("media_type")
    @classmethod
    def image_media_type_is_required(cls, media_type: str) -> str:
        if media_type not in ALLOWED_MEDIA_TYPES:
            raise ValueError("Only image media is permitted")
        return media_type

    @field_validator("filename")
    @classmethod
    def filename_is_present(cls, filename: str) -> str:
        filename = filename.strip()
        if not filename:
            raise ValueError("A filename is required")
        return filename


class CaptureV1Request(CaptureRequest):
    model_config = ConfigDict(extra="forbid", strict=True)

    capture_attempt_key: str = Field(
        min_length=8, max_length=128, pattern=r"^[A-Za-z0-9][A-Za-z0-9_.-]*$"
    )


class ReviewRequest(BaseModel):
    notes: str = Field(min_length=1, max_length=4000)


class SourceContextRequest(BaseModel):
    model_config = ConfigDict(extra="forbid", strict=True)

    evidence_id: str | None = Field(default=None, max_length=80)
    provider: str | None = Field(default=None, max_length=1000)
    origin: str
    source_date: str | None = Field(default=None, max_length=100)
    source_version: str | None = Field(default=None, max_length=500)
    source_location: str | None = Field(default=None, max_length=2000)
    description: str | None = Field(default=None, max_length=4000)
    freshness: str
    limitations: str | None = Field(default=None, max_length=4000)
    duplicate_evidence_id: str | None = Field(default=None, max_length=80)
    source_evidence_ids: list[str] = Field(default_factory=list, max_length=100)
    gap_status: str
    gap_explanation: str | None = Field(default=None, max_length=4000)
    gap_materiality: str | None = Field(default=None, max_length=1000)


class AuditQuestionCreateRequest(BaseModel):
    model_config = ConfigDict(extra="forbid", strict=True)

    engagement_id: str = Field(min_length=1, max_length=80)
    control_id: str = Field(min_length=1, max_length=80)
    question_type: str
    parent_question_id: str | None = Field(default=None, max_length=80)
    question_text: str = Field(min_length=1, max_length=4000)
    purpose: str = Field(min_length=1, max_length=2000)


class AuditQuestionVersionRequest(BaseModel):
    model_config = ConfigDict(extra="forbid", strict=True)

    question_id: str | None = Field(default=None, max_length=80)
    question_text: str = Field(min_length=1, max_length=4000)
    purpose: str = Field(min_length=1, max_length=2000)


class AuditQuestionDecisionRequest(BaseModel):
    model_config = ConfigDict(extra="forbid", strict=True)

    question_id: str | None = Field(default=None, max_length=80)
    decision_attempt_key: str = Field(
        min_length=8, max_length=128, pattern=r"^[A-Za-z0-9][A-Za-z0-9_.-]*$"
    )
    question_version: int = Field(gt=0)
    status: str
    reason: str = Field(min_length=1, max_length=4000)


class ProposedEvidenceLinkRequest(BaseModel):
    model_config = ConfigDict(extra="forbid", strict=True)

    evidence_id: str | None = Field(default=None, max_length=80)
    question_id: str = Field(min_length=1, max_length=80)
    question_version: int = Field(gt=0)
    relevance: str
    reason: str = Field(min_length=1, max_length=4000)


class EvidenceReviewCompletionRequest(BaseModel):
    model_config = ConfigDict(extra="forbid", strict=True)

    evidence_id: str | None = Field(default=None, max_length=80)
    completion_attempt_key: str = Field(
        min_length=8, max_length=128, pattern=r"^[A-Za-z0-9][A-Za-z0-9_.-]*$"
    )
    notes: str = Field(default="Evidence review completed.", min_length=1, max_length=4000)

    @field_validator("notes", mode="before")
    @classmethod
    def default_completion_notes(cls, notes: object) -> object:
        return "Evidence review completed." if notes is None else notes


class EngagementCreateRequest(BaseModel):
    """Strict, safe input for one controlled Engagement draft."""

    model_config = ConfigDict(extra="forbid", strict=True)

    creation_attempt_key: str = Field(
        min_length=8, max_length=128, pattern=r"^[A-Za-z0-9][A-Za-z0-9_.-]*$"
    )
    title: str | None = Field(default=None, max_length=240)
    reference: str | None = Field(default=None, max_length=120)
    authority: str | None = Field(default=None, max_length=2000)
    purpose: str | None = Field(default=None, max_length=2000)
    scope: str | None = Field(default=None, max_length=4000)
    exclusions: str | None = Field(default=None, max_length=4000)
    review_start_date: str | None = Field(default=None, pattern=r"^\d{4}-\d{2}-\d{2}$")
    review_end_date: str | None = Field(default=None, pattern=r"^\d{4}-\d{2}-\d{2}$")
    evidence_cut_off_date: str | None = Field(default=None, pattern=r"^\d{4}-\d{2}-\d{2}$")
    accountable_auditor: str | None = Field(default=None, max_length=240)
    data_classification: str | None = None
    is_fictional: bool | None = None

    @field_validator("data_classification")
    @classmethod
    def data_classification_is_controlled(cls, value: str | None) -> str | None:
        if value is not None and value not in {
            "FICTIONAL",
            "PUBLIC",
            "AUDITCO_OWNED",
            "REAL_CLIENT",
        }:
            raise ValueError("Use a controlled data classification")
        return value


class EngagementActivationRequest(BaseModel):
    model_config = ConfigDict(extra="forbid", strict=True)

    confirmed: bool


class RelationshipReviewDraftRequest(BaseModel):
    model_config = ConfigDict(extra="forbid", strict=True)

    relationship_version: int = Field(gt=0)
    proposed_decision: str | None = Field(
        default=None, pattern="^(APPROVED|REJECTED|CHANGES_REQUIRED)$"
    )
    draft_reason: str | None = Field(default=None, min_length=1, max_length=4000)


class RelationshipApprovalPreviewRequest(BaseModel):
    model_config = ConfigDict(extra="forbid", strict=True)

    relationship_version: int = Field(gt=0)
    approval_reason: str = Field(min_length=1, max_length=4000)


class RelationshipReviewDecisionRequest(BaseModel):
    model_config = ConfigDict(extra="forbid", strict=True)

    relationship_version: int = Field(gt=0)
    decision: str = Field(pattern="^(APPROVED|REJECTED|CHANGES_REQUIRED)$")
    reason: str = Field(min_length=1, max_length=4000)
    decision_key: str = Field(
        min_length=8, max_length=128, pattern=r"^[A-Za-z0-9][A-Za-z0-9_.-]*$"
    )
    preview_token: str | None = Field(
        default=None, pattern=r"^[a-f0-9]{64}$"
    )


class RelationshipRevisionRequest(BaseModel):
    model_config = ConfigDict(extra="forbid", strict=True)

    prior_relationship_version: int = Field(gt=0)
    revision_key: str = Field(
        min_length=8, max_length=128, pattern=r"^[A-Za-z0-9][A-Za-z0-9_.-]*$"
    )
    rationale: str = Field(min_length=1, max_length=4000)
    supporting_source_ids: list[str] = Field(min_length=1, max_length=100)
    gaps: list[str] = Field(max_length=100)
    contradictions: list[str] = Field(max_length=100)
    duplicate_warnings: list[str] = Field(max_length=100)

    @field_validator(
        "supporting_source_ids", "gaps", "contradictions", "duplicate_warnings"
    )
    @classmethod
    def entries_are_bounded_text(cls, values: list[str]) -> list[str]:
        if any(not isinstance(value, str) or not value.strip() or len(value.strip()) > 4000 for value in values):
            raise ValueError("Revision list entries must be bounded text")
        return values


def require_auditor(
    credentials: HTTPBasicCredentials | None = Security(basic_auth),
) -> str:
    password = os.environ.get("ACE_AUDITOR_PASSWORD")
    if not password:
        raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE)
    if (
        credentials is None
        or not secrets.compare_digest(credentials.username, "auditor")
        or not secrets.compare_digest(credentials.password, password)
    ):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            headers={"WWW-Authenticate": "Basic"},
        )
    return credentials.username


_shared_stores: dict[str, WorkbenchStore] = {}


def store() -> WorkbenchStore:
    """Return a shared WorkbenchStore for the configured data directory.

    Reuses the same instance across requests so caches (e.g. the
    evidence-suggestions cache) survive between calls.
    """
    try:
        key = str(resolve_data_dir())
        if key not in _shared_stores:
            _shared_stores[key] = WorkbenchStore()
        return _shared_stores[key]
    except RuntimeError as error:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR) from error


def engagement_service() -> EngagementService:
    return EngagementService(store())


def evidence_review_service() -> EvidenceReviewService:
    return EvidenceReviewService(store())


def relationship_review_service() -> RelationshipReviewService:
    return RelationshipReviewService(store())


def relationship_auditor(auditor: str = Depends(require_auditor)) -> RelationshipReviewActor:
    return RelationshipReviewActor(RelationshipReviewActorKind.ACCOUNTABLE_AUDITOR, auditor)


def relationship_review_queue_content(auditor: RelationshipReviewActor) -> str:
    try:
        queue_response = relationship_review_service().get_queue(auditor)
        queue = queue_response["queue"]
        if not isinstance(queue, list):
            raise ValueError("Relationship Review queue is invalid")
        if not queue:
            return (
                '<p class="muted" id="relationship-reviews-status" role="status" '
                'aria-live="polite">No Relationship Reviews need review.</p>'
                '<ul id="relationship-reviews"></ul>'
            )
        items = []
        for item in queue:
            if not isinstance(item, dict):
                raise ValueError("Relationship Review queue item is invalid")
            relationship_id = quote(str(item["relationship_id"]), safe="")
            items.append(
                "<li>"
                f"<strong>{html.escape(str(item['title']))}</strong> "
                f"({html.escape(str(item['relationship_type']))}) — "
                f"Engagement {html.escape(str(item['engagement_id']))} — "
                f"Version {html.escape(str(item['current_version']))} — "
                f"{html.escape(str(item['material_risk_priority']))} — "
                f"Waiting since {html.escape(str(item['waiting_since']))} — "
                f'<a href="/workbench/relationship-reviews/{relationship_id}">'
                "Review Relationship</a></li>"
            )
        return (
            '<p class="muted" id="relationship-reviews-status" role="status" '
            'aria-live="polite"></p><ul id="relationship-reviews">'
            f"{''.join(items)}</ul>"
        )
    except Exception:
        logger.exception("Relationship Review queue is not available")
        return (
            '<p class="muted" id="relationship-reviews-status" role="status" '
            'aria-live="polite">Relationship Reviews are not available.</p>'
            '<ul id="relationship-reviews"></ul>'
        )


def engagement_response(record: object) -> dict[str, object]:
    return {"api_version": "v1", "engagement": record.as_dict()}  # type: ignore[union-attr]


def engagement_http_error(error: Exception) -> HTTPException:
    if isinstance(error, EngagementNotFoundServiceError):
        return HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Engagement not found")
    if isinstance(error, EngagementG0Error):
        return HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="G0 blocks this Engagement")
    if isinstance(error, EngagementValidationError):
        return HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail=str(error))
    if isinstance(error, EngagementConflictError):
        return HTTPException(status_code=status.HTTP_409_CONFLICT, detail=str(error))
    return HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR)


def evidence_review_http_error(error: Exception) -> HTTPException:
    if isinstance(error, (EvidenceReviewG0Error, EngagementG0Error)):
        return HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="G0 blocks this Evidence Review")
    if isinstance(error, EvidenceReviewNotFoundError):
        return HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Evidence Review not found")
    if isinstance(error, EvidenceReviewValidationError):
        return HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail=str(error))
    if isinstance(error, EvidenceReviewConflictError):
        return HTTPException(status_code=status.HTTP_409_CONFLICT, detail=str(error))
    return HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR)


def relationship_review_http_error(error: Exception) -> HTTPException:
    if isinstance(error, KeyError):
        return HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Relationship Review item not found")
    if isinstance(error, ValueError):
        return HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail=str(error))
    return HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR)


def relationship_review_response(result: dict[str, object]) -> dict[str, object]:
    return {"api_version": "v1", **result}


def decode_image(payload: CaptureRequest) -> bytes:
    try:
        content = base64.b64decode(payload.data_base64, validate=True)
    except (binascii.Error, ValueError) as error:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="Invalid image data") from error
    if not content or len(content) > MAX_IMAGE_BYTES:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="Image size is invalid")
    try:
        _verify_image(content, IMAGE_FORMATS[payload.media_type])
    except (
        Image.DecompressionBombWarning,
        Image.DecompressionBombError,
        OSError,
        SyntaxError,
        UnidentifiedImageError,
        ValueError,
    ) as error:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="Image data does not match media type")
    return content


def _verify_image(content: bytes, expected_format: str) -> None:
    with warnings.catch_warnings():
        warnings.simplefilter("error", Image.DecompressionBombWarning)
        with Image.open(BytesIO(content)) as image:
            if image.format != expected_format:
                raise ValueError("Image data does not match media type")
            image.verify()
        with Image.open(BytesIO(content)) as image:
            if image.format != expected_format:
                raise ValueError("Image data does not match media type")
            for frame_number in range(getattr(image, "n_frames", 1)):
                image.seek(frame_number)
                image.load()


@router.get("", response_class=HTMLResponse)
def workbench_page(auditor: str = Depends(require_auditor)) -> HTMLResponse:
    try:
        current = engagement_service().current()
    except EngagementG0Error as error:
        raise engagement_http_error(error) from error
    current_text = (
        f"{current.title} — {current.reference} — {current.state}"
        if current is not None
        else "No READY_FOR_CAPTURE Engagement selected"
    )
    relationship_content = relationship_review_queue_content(
        RelationshipReviewActor(RelationshipReviewActorKind.ACCOUNTABLE_AUDITOR, auditor)
    )
    return HTMLResponse(
        WORKBENCH_PAGE.replace("__CURRENT_ENGAGEMENT__", html.escape(current_text)).replace(
            "__RELATIONSHIP_REVIEW_CONTENT__", relationship_content
        )
    )


@router.get("/engagements/new", response_class=HTMLResponse)
def engagement_setup_page(_: str = Depends(require_auditor)) -> HTMLResponse:
    return HTMLResponse(ENGAGEMENT_SETUP_PAGE)


@router.post("/api/v1/engagements", status_code=status.HTTP_201_CREATED)
def create_engagement(
    payload: EngagementCreateRequest, auditor: str = Depends(require_auditor)
) -> dict[str, object]:
    try:
        record = engagement_service().create_draft(EngagementDraft(**payload.model_dump()), auditor)
    except (EngagementConflictError, EngagementG0Error) as error:
        raise engagement_http_error(error) from error
    return engagement_response(record)


@router.get("/api/v1/engagements/current")
def current_engagement(_: str = Depends(require_auditor)) -> dict[str, object]:
    try:
        record = engagement_service().current()
    except EngagementG0Error as error:
        raise engagement_http_error(error) from error
    if record is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No current Engagement")
    return engagement_response(record)


@router.get("/api/v1/engagements/{engagement_id}")
def get_engagement(
    engagement_id: str, _: str = Depends(require_auditor)
) -> dict[str, object]:
    try:
        return engagement_response(engagement_service().get(engagement_id))
    except (EngagementG0Error, EngagementNotFoundServiceError) as error:
        raise engagement_http_error(error) from error


@router.post("/api/v1/engagements/{engagement_id}/activate")
def activate_engagement(
    engagement_id: str,
    payload: EngagementActivationRequest,
    auditor: str = Depends(require_auditor),
) -> dict[str, object]:
    try:
        return engagement_response(
            engagement_service().activate(engagement_id, payload.confirmed, auditor)
        )
    except (
        EngagementConflictError,
        EngagementG0Error,
        EngagementNotFoundServiceError,
        EngagementValidationError,
    ) as error:
        raise engagement_http_error(error) from error


@router.post("/api/v1/engagements/sync", status_code=status.HTTP_201_CREATED)
def sync_engagement(
    payload: EngagementCreateRequest, auditor: str = Depends(require_auditor)
) -> dict[str, object]:
    """Accept an offline draft. Idempotent: same creation_attempt_key returns existing."""
    try:
        record = engagement_service().create_draft(
            EngagementDraft(**payload.model_dump()), auditor
        )
    except (EngagementConflictError, EngagementG0Error) as error:
        raise engagement_http_error(error) from error
    response = engagement_response(record)
    response["synced"] = True
    return response


@router.put("/api/v1/engagements/{engagement_id}/current")
def select_current_engagement(
    engagement_id: str, _: str = Depends(require_auditor)
) -> dict[str, object]:
    try:
        return engagement_response(engagement_service().select_current(engagement_id))
    except (EngagementConflictError, EngagementG0Error, EngagementNotFoundServiceError) as error:
        raise engagement_http_error(error) from error


@router.get("/summary")
def workbench_summary(_: str = Depends(require_auditor)) -> dict[str, object]:
    workbench_store = store()
    try:
        current = EngagementService(workbench_store).current()
    except EngagementG0Error as error:
        raise engagement_http_error(error) from error
    summary = workbench_store.summary()
    summary["current_engagement"] = current.as_dict() if current is not None else None
    summary["engagement"] = current.title if current is not None else "No current Engagement"
    return summary


@router.get("/engagement/summary", response_class=HTMLResponse)
def engagement_control_summary(_: str = Depends(require_auditor)) -> HTMLResponse:
    workbench_store = store()
    try:
        EngagementService(workbench_store).current()
    except EngagementG0Error as error:
        raise engagement_http_error(error) from error
    data = workbench_store.engagement_summary()
    html_content = _render_engagement_summary_html(data)
    return HTMLResponse(html_content)


def _render_engagement_summary_html(data: dict[str, object]) -> str:
    if not data.get("has_engagement"):
        return ENGAGEMENT_SUMMARY_PAGE.replace(
            "__CONTENT__",
            '<section><h2>No Current Engagement</h2>'
            '<p class="muted">Set up an Engagement to see the control summary.</p>'
            '<p><a href="/workbench/engagements/new">Create An Engagement</a></p></section>',
        )
    eng: dict[str, object] = data["engagement"]  # type: ignore[assignment]
    evidence: dict[str, object] = data["evidence"]  # type: ignore[assignment]
    recent: list[dict[str, str]] = data["recent_events"]  # type: ignore[assignment]
    recommendation: str = data["recommendation"]  # type: ignore[assignment]
    open_conflicts: int = data["open_conflicts"]  # type: ignore[assignment]
    open_gaps: int = data["open_gaps"]  # type: ignore[assignment]
    pending_items: list[dict[str, str]] = data.get("pending_items", [])  # type: ignore[assignment]
    pending_review_count = evidence.get("pending_review", 0) if isinstance(evidence.get("pending_review"), int) else 0

    def esc(value: object) -> str:
        return html.escape(str(value)) if value is not None else "—"

    def date_esc(value: object) -> str:
        if value is None:
            return "—"
        try:
            from datetime import date as date_type

            date_type.fromisoformat(str(value))
            return str(value)
        except (ValueError, TypeError):
            return html.escape(str(value))

    rows: list[str] = []
    rows.append(f"<tr><th>Title</th><td>{esc(eng.get('title'))}</td></tr>")
    rows.append(f"<tr><th>Reference</th><td>{esc(eng.get('reference'))}</td></tr>")
    rows.append(f"<tr><th>State</th><td>{esc(eng.get('state'))}</td></tr>")
    rows.append(f"<tr><th>Data Classification</th><td>{esc(eng.get('data_classification'))}</td></tr>")
    rows.append(f"<tr><th>Accountable Auditor</th><td>{esc(eng.get('accountable_auditor'))}</td></tr>")
    rows.append(f"<tr><th>Authority</th><td>{esc(eng.get('authority'))}</td></tr>")
    rows.append(f"<tr><th>Purpose</th><td>{esc(eng.get('purpose'))}</td></tr>")
    rows.append(f"<tr><th>Scope</th><td>{esc(eng.get('scope'))}</td></tr>")
    rows.append(f"<tr><th>Exclusions</th><td>{esc(eng.get('exclusions'))}</td></tr>")
    rows.append(f"<tr><th>Review Start Date</th><td>{date_esc(eng.get('review_start_date'))}</td></tr>")
    rows.append(f"<tr><th>Review End Date</th><td>{date_esc(eng.get('review_end_date'))}</td></tr>")
    rows.append(f"<tr><th>Evidence Cut-Off Date</th><td>{date_esc(eng.get('evidence_cut_off_date'))}</td></tr>")

    evidence_rows = [
        f"<tr><td>Evidence Captured</td><td>{esc(evidence.get('captured'))}</td></tr>",
        f"<tr><td>Pending Review</td><td>{esc(evidence.get('pending_review'))}</td></tr>",
        f"<tr><td>Evidence Reviewed</td><td>{esc(evidence.get('reviewed'))}</td></tr>",
    ]

    open_rows: list[str] = []
    if open_conflicts > 0:
        open_rows.append(f"<tr><td>Open Relationship Conflicts</td><td>{open_conflicts}</td></tr>")
    if open_gaps > 0:
        open_rows.append(f"<tr><td>Open Gaps (controlled source)</td><td>{open_gaps}</td></tr>")
    if not open_rows:
        open_rows.append('<tr><td colspan="2" class="muted">No open conflicts or gaps.</td></tr>')
    open_rows.append(
        '<tr><td colspan="2" class="muted">'
        'Relationship conflicts appear only for linked captured evidence. '
        'Trace-level conflicts are managed in '
        '<a href="/workbench/relationship-reviews">Relationship Review</a>.'
        '</td></tr>'
    )

    recent_items: list[str] = []
    for event in recent:
        recent_items.append(
            f"<li>{esc(event.get('event_type'))} — "
            f"{esc(event.get('actor'))} — "
            f"{esc(event.get('recorded_at'))}</li>"
        )
    if not recent_items:
        recent_items.append('<li class="muted">No recent activity recorded.</li>')

    pending_preview: list[str] = []
    if pending_items:
        pending_preview.append('<table><tbody>')
        for item in pending_items[:5]:
            evidence_id = esc(item.get("evidence_id"))
            filename = esc(item.get("filename"))
            captured = esc(item.get("captured_at"))
            pending_preview.append(
                f'<tr><td><a href="/workbench/evidence/{evidence_id}">'
                f'{filename}</a></td><td>{captured}</td></tr>'
            )
        pending_preview.append('</tbody></table>')
    else:
        pending_preview.append('<p class="muted">No evidence pending review.</p>')

    rec_html = esc(recommendation)
    if pending_review_count > 0:
        rec_html = (
            f'{esc(recommendation)} '
            f'<a href="/workbench">Open the Field Capture Workbench to review.</a>'
        )
    elif open_conflicts > 0:
        rec_html = (
            f'{esc(recommendation)} '
            f'<a href="/workbench/relationship-reviews">Open Relationship Review.</a>'
        )

    content = (
        '<section><h2>Engagement</h2><table><tbody>'
        f"{''.join(rows)}</tbody></table></section>"
        '<section><h2>Evidence</h2><table><tbody>'
        f"{''.join(evidence_rows)}</tbody></table></section>"
        '<section><h2>Pending Evidence</h2>'
        f"{''.join(pending_preview)}</section>"
        '<section><h2>Open Items</h2><table><tbody>'
        f"{''.join(open_rows)}</tbody></table></section>"
        '<section><h2>Recent Activity</h2><ul>'
        f"{''.join(recent_items)}</ul></section>"
        '<section><h2>Recommended Next Action</h2>'
        f"<p>{rec_html}</p></section>"
    )
    return ENGAGEMENT_SUMMARY_PAGE.replace("__CONTENT__", content)


@router.get("/engagement/graph", response_class=HTMLResponse)
def engagement_graph(_: str = Depends(require_auditor)) -> HTMLResponse:
    """Read-only graph projection of the current fictional engagement."""
    workbench_store = store()
    try:
        EngagementService(workbench_store).current()
    except EngagementG0Error as error:
        raise engagement_http_error(error) from error
    data = workbench_store.graph_projection()
    return HTMLResponse(_render_graph_page(data))


@router.get("/engagement/graph/export")
def engagement_graph_export(_: str = Depends(require_auditor)) -> "Response":
    """Download an XLSX workbook of the current engagement graph."""
    workbench_store = store()
    try:
        EngagementService(workbench_store).current()
    except EngagementG0Error as error:
        raise engagement_http_error(error) from error
    data = workbench_store.graph_projection()
    return _build_graph_xlsx(data)


# ---------------------------------------------------------------------------
# Graph HTML rendering (inline SVG — no external libraries)
# ---------------------------------------------------------------------------

_GRAPH_PAGE = """<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Engagement Graph — ACE Workbench</title>
<style>
  :root { color-scheme: light dark; }
  body { font-family: system-ui, sans-serif; max-width: 960px; margin: 2rem auto; padding: 0 1rem; }
  h1 { margin-bottom: 0.25rem; }
  .muted { color: #6b7280; font-size: 0.875rem; }
  table { width: 100%; border-collapse: collapse; margin: 1.5rem 0; }
  th, td { text-align: left; padding: 0.5rem 0.75rem; border-bottom: 1px solid #e5e7eb; }
  th { font-weight: 600; background: #f9fafb; }
  .warn { background: #fef3c7; }
  .conflict { background: #fee2e2; }
  .badge { display: inline-block; padding: 0.125rem 0.5rem; border-radius: 999px; font-size: 0.75rem; font-weight: 600; }
  .badge-fact { background: #d1fae5; color: #065f46; }
  .badge-warn { background: #fef3c7; color: #92400e; }
  .badge-conflict { background: #fee2e2; color: #991b1b; }
  svg { display: block; margin: 2rem auto; max-width: 100%; }
  .node-box { fill: #f0f9ff; stroke: #0284c7; stroke-width: 1.5; rx: 6; }
  .node-box-evidence { fill: #ecfdf5; stroke: #059669; }
  .node-box-conflict { fill: #fef2f2; stroke: #dc2626; }
  .node-text { font-size: 11px; fill: #1f2937; }
  .node-type { font-size: 9px; fill: #6b7280; }
  .edge-line { stroke: #9ca3af; stroke-width: 1.5; marker-end: url(#arrowhead); }
  .edge-line-conflict { stroke: #dc2626; stroke-dasharray: 5,3; marker-end: url(#arrowhead-conflict); }
  .edge-label { font-size: 9px; fill: #6b7280; }
  .empty-state { text-align: center; padding: 4rem 2rem; }
  .export-link { float: right; }
  a { color: #2563eb; }
</style>
</head>
<body>
<h1>Engagement Graph <a class="export-link" href="/workbench/engagement/graph/export">⬇ Export XLSX</a></h1>
<p class="muted">Read-only. Nodes and edges from the current fictional engagement.</p>
__GRAPH_SVG__
__GRAPH_TABLE__
__GRAPH_RELS__
__GRAPH_WARNINGS__
__GRAPH_EMPTY__
</body>
</html>"""


def _render_graph_page(data: dict[str, object]) -> str:
    nodes: list[dict[str, object]] = data.get("nodes", [])  # type: ignore[assignment]
    edges: list[dict[str, object]] = data.get("edges", [])  # type: ignore[assignment]
    warnings: list[dict[str, object]] = data.get("warnings", [])  # type: ignore[assignment]
    engagement_id = data.get("engagement_id")

    if engagement_id is None:
        return _GRAPH_PAGE.replace("__GRAPH_SVG__", "").replace("__GRAPH_TABLE__", "").replace(
            "__GRAPH_RELS__", ""
        ).replace("__GRAPH_WARNINGS__", "").replace(
            "__GRAPH_EMPTY__",
            '<div class="empty-state"><h2>No current engagement</h2><p class="muted">Activate a fictional engagement to see its graph.</p></div>',
        )
    if not nodes:
        return _GRAPH_PAGE.replace("__GRAPH_SVG__", "").replace("__GRAPH_TABLE__", "").replace(
            "__GRAPH_RELS__", ""
        ).replace("__GRAPH_WARNINGS__", "").replace(
            "__GRAPH_EMPTY__",
            '<div class="empty-state"><h2>No records found</h2><p class="muted">The current engagement has no approved records to graph.</p></div>',
        )

    # --- Inline SVG ---
    conflict_edge_ids = {f"{e['source']}→{e['target']}" for e in edges if e.get("status") == "OPEN" or e.get("type") == "CONTRA"}
    svg_width = 500
    node_h = 36
    gap = 18
    total_h = len(nodes) * (node_h + gap) + 40

    svg_parts: list[str] = [
        f'<svg viewBox="0 0 {svg_width} {total_h}" xmlns="http://www.w3.org/2000/svg">',
        '<defs><marker id="arrowhead" markerWidth="8" markerHeight="6" refX="8" refY="3" orient="auto"><polygon points="0 0, 8 3, 0 6" fill="#9ca3af"/></marker>',
        '<marker id="arrowhead-conflict" markerWidth="8" markerHeight="6" refX="8" refY="3" orient="auto"><polygon points="0 0, 8 3, 0 6" fill="#dc2626"/></marker></defs>',
    ]

    # Build node-id → y-position map for edge rendering
    node_y: dict[str, int] = {}
    for idx, node in enumerate(nodes):
        y = 20 + idx * (node_h + gap)
        ntype = str(node.get("type", ""))
        nid = str(node["id"])
        node_y[nid] = y
        is_evidence = ntype == "evidence"
        has_conflict = any(nid in c for c in conflict_edge_ids)
        box_class = "node-box-conflict" if has_conflict else ("node-box-evidence" if is_evidence else "node-box")
        svg_parts.append(f'<rect x="40" y="{y}" width="420" height="{node_h}" class="{box_class}"/>')
        svg_parts.append(f'<text x="50" y="{y + 14}" class="node-type">{ntype}</text>')
        label = str(node.get("label", nid))[:60]
        svg_parts.append(f'<text x="50" y="{y + 28}" class="node-text">{_esc(label)}</text>')
        svg_parts.append(f'<text x="460" y="{y + 28}" class="node-type" text-anchor="end">{nid}</text>')

    # Draw arrows for all edges between nodes that both appear in the graph
    for edge in edges:
        source_id = str(edge["source"])
        target_id = str(edge["target"])
        if source_id not in node_y or target_id not in node_y:
            continue
        y_src = node_y[source_id]
        y_tgt = node_y[target_id]
        edge_key = f"{source_id}→{target_id}"
        # Draw line from bottom of source to top of target
        arrow_y1 = y_src + node_h + 2
        arrow_y2 = y_tgt - 4
        if arrow_y2 <= arrow_y1:
            arrow_y2 = arrow_y1 + gap - 4
        line_class = "edge-line-conflict" if edge_key in conflict_edge_ids else "edge-line"
        svg_parts.append(f'<line x1="250" y1="{arrow_y1}" x2="250" y2="{arrow_y2}" class="{line_class}" marker-end="url(#arrowhead)"/>')
        etype = str(edge.get("type", ""))
        estatus = str(edge.get("status", ""))
        mid_y = (arrow_y1 + arrow_y2) // 2
        svg_parts.append(f'<text x="255" y="{mid_y}" class="edge-label">{_esc(etype)} ({_esc(estatus)})</text>')

    svg_parts.append("</svg>")

    # --- Table ---
    table_rows: list[str] = []
    for node in nodes:
        ntype = str(node.get("type", ""))
        nid = str(node["id"])
        status = str(node.get("status", "")) if node.get("status") else "—"
        row_class = ""
        if ntype == "evidence" and bool(node.get("is_capture")) is False:
            row_class = ' class="warn"'
        table_rows.append(
            f'<tr{row_class}><td><a href="/workbench/evidence/{_esc(nid)}">{_esc(nid)}</a></td>'
            f"<td>{_esc(ntype)}</td><td>{_esc(str(node.get('label', ''))[:80])}</td>"
            f"<td>{_esc(status)}</td></tr>"
        )

    # --- Warnings section ---
    warn_html = ""
    if warnings:
        warn_items = []
        for w in warnings:
            level = str(w.get("level", "INFO"))
            cls = "conflict" if level == "WARNING" else "warn"
            warn_items.append(
                f'<tr class="{cls}"><td>{_esc(str(w.get("record_id", "")))}</td>'
                f"<td>{_esc(str(w.get('detail', '')))}</td></tr>"
            )
        warn_html = "<section><h2>Warnings</h2><table><thead><tr><th>Record</th><th>Detail</th></tr></thead><tbody>" + "".join(warn_items) + "</tbody></table></section>"

    table_html = '<section><h2>Records</h2><table><thead><tr><th>ID</th><th>Type</th><th>Label</th><th>Status</th></tr></thead><tbody>' + "".join(table_rows) + "</tbody></table></section>"

    # --- Relationships direction table ---
    rel_rows: list[str] = []
    for edge in edges:
        rel_rows.append(
            f"<tr><td>{_esc(str(edge['source']))} → {_esc(str(edge['target']))}</td>"
            f"<td>{_esc(str(edge.get('type', '')))}</td>"
            f"<td>{_esc(str(edge.get('status', '')))}</td></tr>"
        )
    rel_table_html = '<section><h2>Relationships</h2><table><thead><tr><th>Direction</th><th>Type</th><th>Status</th></tr></thead><tbody>' + "".join(rel_rows) + "</tbody></table></section>"

    return (
        _GRAPH_PAGE.replace("__GRAPH_SVG__", "".join(svg_parts))
        .replace("__GRAPH_TABLE__", table_html)
        .replace("__GRAPH_RELS__", rel_table_html)
        .replace("__GRAPH_WARNINGS__", warn_html)
        .replace("__GRAPH_EMPTY__", "")
    )


def _esc(text: str) -> str:
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def _xlsx_safe(value: object) -> str:
    """Escape formula-injection prefixes in XLSX cell values.

    Prepends a single-quote prefix to values that start with ``=``,
    ``+``, ``-``, or ``@`` so spreadsheet applications treat them as
    literal text rather than executable formulas.
    """
    s = str(value) if value is not None else ""
    if s and s[0] in ("=", "+", "-", "@"):
        return "'" + s
    return s


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------


def _build_graph_xlsx(data: dict[str, object]) -> "Response":
    from datetime import datetime, timezone

    from openpyxl import Workbook
    from starlette.responses import Response as StarletteResponse

    nodes: list[dict[str, object]] = data.get("nodes", [])  # type: ignore[assignment]
    edges: list[dict[str, object]] = data.get("edges", [])  # type: ignore[assignment]
    warnings: list[dict[str, object]] = data.get("warnings", [])  # type: ignore[assignment]
    engagement_id = str(data.get("engagement_id", "unknown"))
    export_time = datetime.now(timezone.utc).isoformat()
    export_id = f"{engagement_id}-{export_time}"

    wb = Workbook()

    # --- Records sheet ---
    ws_records = wb.active
    ws_records.title = "Records"
    ws_records.append(["ID", "Type", "Label", "Status"])
    for n in nodes:
        ws_records.append([
            _xlsx_safe(n["id"]),
            _xlsx_safe(n["type"]),
            _xlsx_safe(n.get("label", "")),
            _xlsx_safe(n.get("status", "")),
        ])

    # --- Evidence sheet ---
    ws_evidence = wb.create_sheet("Evidence")
    ws_evidence.append(["Evidence ID", "Filename", "Media Type", "Photo Status", "Has Source Text"])
    for n in nodes:
        if n.get("type") != "evidence":
            continue
        media = n.get("media_type") or "—"
        photo_status = "Not captured" if not bool(n.get("is_capture")) else "Captured"
        has_source = "Yes" if n.get("has_source_text") else "No"
        ws_evidence.append([
            _xlsx_safe(n["id"]),
            _xlsx_safe(n.get("label", "")),
            _xlsx_safe(media),
            _xlsx_safe(photo_status),
            _xlsx_safe(has_source),
        ])

    # --- Relationships sheet ---
    ws_rels = wb.create_sheet("Relationships")
    ws_rels.append(["Source ID", "Target ID", "Type", "Status"])
    for e in edges:
        ws_rels.append([
            _xlsx_safe(e["source"]),
            _xlsx_safe(e["target"]),
            _xlsx_safe(e["type"]),
            _xlsx_safe(e["status"]),
        ])

    # --- Warnings sheet ---
    ws_warn = wb.create_sheet("Warnings")
    ws_warn.append(["Record ID", "Level", "Detail"])
    for w in warnings:
        ws_warn.append([
            _xlsx_safe(w.get("record_id", "")),
            _xlsx_safe(w.get("level", "")),
            _xlsx_safe(w.get("detail", "")),
        ])

    # --- Read Me sheet ---
    ws_readme = wb.create_sheet("Read Me")
    ws_readme.append(["Export ID", _xlsx_safe(export_id)])
    ws_readme.append(["Engagement ID", _xlsx_safe(engagement_id)])
    ws_readme.append(["Export Time (UTC)", _xlsx_safe(export_time)])
    ws_readme.append(["Notes", "Read-only review copy. Does not write back to ACE SQLite."])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StarletteResponse(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="ace-graph-{engagement_id}.xlsx"'},
    )


@router.post("/evidence", status_code=status.HTTP_201_CREATED)
def capture_evidence(
    payload: CaptureRequest, auditor: str = Depends(require_auditor)
) -> dict[str, str]:
    workbench_store = store()
    try:
        return EngagementService(workbench_store).capture(
            payload.filename,
            payload.media_type,
            decode_image(payload),
            auditor,
            payload.capture_attempt_key,
        )
    except EngagementG0Error as error:
        raise engagement_http_error(error) from error
    except (CaptureAttemptConflictError, EngagementConflictError, NoReadyCurrentEngagementError) as error:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "Capture attempt key conflicts"
                if isinstance(error, CaptureAttemptConflictError)
                else "Select a READY_FOR_CAPTURE Engagement before capture"
            ),
        ) from error


@router.post("/api/v1/evidence", status_code=status.HTTP_201_CREATED)
def capture_evidence_v1(
    payload: CaptureV1Request, auditor: str = Depends(require_auditor)
) -> dict[str, object]:
    workbench_store = store()
    try:
        evidence = EngagementService(workbench_store).capture(
            payload.filename,
            payload.media_type,
            decode_image(payload),
            auditor,
            payload.capture_attempt_key,
        )
    except EngagementG0Error as error:
        raise engagement_http_error(error) from error
    except (CaptureAttemptConflictError, EngagementConflictError, NoReadyCurrentEngagementError) as error:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "Capture attempt key conflicts"
                if isinstance(error, CaptureAttemptConflictError)
                else "Select a READY_FOR_CAPTURE Engagement before capture"
            ),
        ) from error
    return {"api_version": "v1", "evidence": evidence}


@router.get("/evidence/{evidence_id}/media")
def evidence_media(evidence_id: str, _: str = Depends(require_auditor)) -> FileResponse:
    try:
        media = store().media(evidence_id)
    except (EvidenceReviewConflictError, EvidenceReviewG0Error, EvidenceReviewNotFoundError, EvidenceReviewValidationError) as error:
        raise evidence_review_http_error(error) from error
    except ValueError as error:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND) from error
    if media is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND)
    file_path, media_type = media
    return FileResponse(file_path, media_type=media_type, filename=file_path.name)


@router.get("/evidence/{evidence_id}/review", response_class=HTMLResponse)
def evidence_review_page(evidence_id: str, _: str = Depends(require_auditor)) -> HTMLResponse:
    try:
        evidence_review_service().get(evidence_id)
    except (EvidenceReviewConflictError, EvidenceReviewG0Error, EvidenceReviewNotFoundError, EvidenceReviewValidationError) as error:
        raise evidence_review_http_error(error) from error
    return HTMLResponse(EVIDENCE_REVIEW_PAGE.replace("__EVIDENCE_ID__", html.escape(evidence_id)))


@router.get("/api/v1/evidence/{evidence_id}/review")
def get_evidence_review(
    evidence_id: str, _: str = Depends(require_auditor)
) -> dict[str, object]:
    try:
        return {"api_version": "v1", **evidence_review_service().get(evidence_id)}
    except (EvidenceReviewConflictError, EvidenceReviewG0Error, EvidenceReviewNotFoundError, EvidenceReviewValidationError) as error:
        raise evidence_review_http_error(error) from error


@router.get("/api/v1/evidence/{evidence_id}/suggestions")
def get_evidence_suggestions(
    evidence_id: str, _: str = Depends(require_auditor)
) -> dict[str, object]:
    return {"api_version": "v1", **store().evidence_suggestions(evidence_id)}


@router.put("/api/v1/evidence/{evidence_id}/review/context")
def save_evidence_review_context(
    evidence_id: str, payload: SourceContextRequest, auditor: str = Depends(require_auditor)
) -> dict[str, object]:
    try:
        state = evidence_review_service().save_context(evidence_id, payload.model_dump(), auditor)
        return {"api_version": "v1", **state}
    except (EvidenceReviewConflictError, EvidenceReviewG0Error, EvidenceReviewNotFoundError, EvidenceReviewValidationError) as error:
        raise evidence_review_http_error(error) from error


@router.post("/api/v1/audit-questions", status_code=status.HTTP_201_CREATED)
def create_audit_question(
    payload: AuditQuestionCreateRequest, auditor: str = Depends(require_auditor)
) -> dict[str, object]:
    try:
        return {"api_version": "v1", "question_version": evidence_review_service().create_question(payload.model_dump(), auditor)}
    except (EvidenceReviewConflictError, EvidenceReviewG0Error, EvidenceReviewNotFoundError, EvidenceReviewValidationError) as error:
        raise evidence_review_http_error(error) from error


@router.post("/api/v1/audit-questions/{question_id}/versions", status_code=status.HTTP_201_CREATED)
def create_audit_question_version(
    question_id: str, payload: AuditQuestionVersionRequest, auditor: str = Depends(require_auditor)
) -> dict[str, object]:
    try:
        result = evidence_review_service().create_question_version(question_id, payload.model_dump(), auditor)
        return {"api_version": "v1", "question_version": result}
    except (EvidenceReviewConflictError, EvidenceReviewG0Error, EvidenceReviewNotFoundError, EvidenceReviewValidationError) as error:
        raise evidence_review_http_error(error) from error


@router.post("/api/v1/audit-questions/{question_id}/decisions", status_code=status.HTTP_201_CREATED)
def decide_audit_question(
    question_id: str, payload: AuditQuestionDecisionRequest, auditor: str = Depends(require_auditor)
) -> dict[str, object]:
    try:
        result = evidence_review_service().decide_question(question_id, payload.model_dump(), auditor)
        return {"api_version": "v1", "decision": result}
    except (EvidenceReviewConflictError, EvidenceReviewG0Error, EvidenceReviewNotFoundError, EvidenceReviewValidationError) as error:
        raise evidence_review_http_error(error) from error


@router.post("/api/v1/evidence/{evidence_id}/proposed-links", status_code=status.HTTP_201_CREATED)
def create_proposed_evidence_link(
    evidence_id: str, payload: ProposedEvidenceLinkRequest, auditor: str = Depends(require_auditor)
) -> dict[str, object]:
    try:
        result = evidence_review_service().propose_link(evidence_id, payload.model_dump(), auditor)
        return {"api_version": "v1", "proposed_link": result}
    except (EvidenceReviewConflictError, EvidenceReviewG0Error, EvidenceReviewNotFoundError, EvidenceReviewValidationError) as error:
        raise evidence_review_http_error(error) from error


@router.post("/api/v1/evidence/{evidence_id}/review/complete")
def complete_evidence_review(
    evidence_id: str, payload: EvidenceReviewCompletionRequest, auditor: str = Depends(require_auditor)
) -> dict[str, object]:
    try:
        result = evidence_review_service().complete(evidence_id, payload.model_dump(), auditor)
        return {"api_version": "v1", "completion": result}
    except (
        EvidenceReviewConflictError,
        EvidenceReviewG0Error,
        EvidenceReviewNotFoundError,
        EvidenceReviewValidationError,
    ) as error:
        raise evidence_review_http_error(error) from error


@router.post("/evidence/{evidence_id}/review")
def review_evidence(
    evidence_id: str, payload: ReviewRequest, auditor: str = Depends(require_auditor)
) -> dict[str, str]:
    try:
        return evidence_review_service().complete_legacy(evidence_id, payload.notes, auditor)  # type: ignore[return-value]
    except (
        EvidenceReviewConflictError,
        EvidenceReviewG0Error,
        EvidenceReviewNotFoundError,
        EvidenceReviewValidationError,
    ) as error:
        raise evidence_review_http_error(error) from error


@router.get("/relationship-reviews/{relationship_id}", response_class=HTMLResponse)
def relationship_review_page(
    relationship_id: str, auditor: RelationshipReviewActor = Depends(relationship_auditor)
) -> HTMLResponse:
    try:
        response = relationship_review_service().get_item(relationship_id, auditor)
    except (KeyError, ValueError) as error:
        raise relationship_review_http_error(error) from error
    rendered_state = html.escape(
        json.dumps(relationship_review_response(response), sort_keys=True)
    )
    if "state" not in response:
        result = response["result"]
        content = f"<p role=\"status\">{html.escape(str(result['message']))}</p>"
    else:
        state = response["state"]
        source_support = "".join(
            f"<li>{html.escape(str(source_id))}</li>"
            for source_id in state["source_support"]
        )
        warnings = "".join(
            f"<li>{html.escape(label)}: {html.escape(str(value))}</li>"
            for label, values in (
                ("Gap", state["gaps"]),
                ("Contradiction", state["contradictions"]),
                ("Duplicate Warning", state["duplicate_warnings"]),
            )
            for value in values
        )
        versions = "".join(
            f"<li>Version {version['version']}: {html.escape(str(version['rationale']))}</li>"
            for version in state["version_history"]
        )
        decisions = "".join(
            f"<li>Version {decision['relationship_version']}: "
            f"{html.escape(str(decision['decision_status']))} — "
            f"{html.escape(str(decision['reason']))}</li>"
            for decision in state["decisions"]
        ) or "<li>None recorded.</li>"
        content = (
            "<section><h2>Linked Records</h2><dl>"
            f"<dt>Source Record</dt><dd>{html.escape(str(state['source_record_id']))}</dd>"
            f"<dt>Target Record</dt><dd>{html.escape(str(state['target_record_id']))}</dd>"
            f"<dt>Relationship Type</dt><dd>{html.escape(str(state['relationship_type']))}</dd>"
            "</dl></section>"
            "<section><h2>Rationale</h2>"
            f"<p>{html.escape(str(state['rationale']))}</p></section>"
            f"<section><h2>Source Support</h2><ul>{source_support}</ul></section>"
            f"<section><h2>Warnings</h2><ul>{warnings}</ul></section>"
            f"<section><h2>Version History</h2><ul>{versions}</ul></section>"
            f"<section><h2>Earlier Decisions</h2><ul>{decisions}</ul></section>"
        )
    return HTMLResponse(
        "<!doctype html><html lang=\"en-AU\"><head><meta charset=\"utf-8\">"
        "<title>Relationship Review | ACE Workbench</title></head><body>"
        "<main><h1>Relationship Review</h1><p>Accountable Auditor final-decision control applies.</p>"
        f"{content}<pre id=\"relationship-review-state\">{rendered_state}</pre>"
        "</main></body></html>"
    )


@router.get("/api/v1/relationship-reviews")
def relationship_review_queue(
    auditor: RelationshipReviewActor = Depends(relationship_auditor),
) -> dict[str, object]:
    try:
        return relationship_review_response(relationship_review_service().get_queue(auditor))
    except (KeyError, ValueError) as error:
        raise relationship_review_http_error(error) from error


@router.get("/api/v1/relationship-reviews/{relationship_id}")
def relationship_review_item(
    relationship_id: str, auditor: RelationshipReviewActor = Depends(relationship_auditor)
) -> dict[str, object]:
    try:
        return relationship_review_response(relationship_review_service().get_item(relationship_id, auditor))
    except (KeyError, ValueError) as error:
        raise relationship_review_http_error(error) from error


@router.put("/api/v1/relationship-reviews/{relationship_id}/draft")
def save_relationship_review_draft(
    relationship_id: str,
    payload: RelationshipReviewDraftRequest,
    auditor: RelationshipReviewActor = Depends(relationship_auditor),
) -> dict[str, object]:
    try:
        return relationship_review_response(
            relationship_review_service().save_draft(relationship_id, payload.model_dump(), auditor)
        )
    except (KeyError, ValueError) as error:
        raise relationship_review_http_error(error) from error


@router.post("/api/v1/relationship-reviews/{relationship_id}/approval-preview")
def relationship_approval_preview(
    relationship_id: str,
    payload: RelationshipApprovalPreviewRequest,
    auditor: RelationshipReviewActor = Depends(relationship_auditor),
) -> dict[str, object]:
    try:
        return relationship_review_response(
            relationship_review_service().request_approval_preview(
                relationship_id, payload.model_dump(), auditor
            )
        )
    except (KeyError, ValueError) as error:
        raise relationship_review_http_error(error) from error


@router.post("/api/v1/relationship-reviews/{relationship_id}/decision")
def relationship_final_decision(
    relationship_id: str,
    payload: RelationshipReviewDecisionRequest,
    auditor: RelationshipReviewActor = Depends(relationship_auditor),
) -> dict[str, object]:
    try:
        return relationship_review_response(
            relationship_review_service().record_decision(relationship_id, payload.model_dump(), auditor)
        )
    except (KeyError, ValueError) as error:
        raise relationship_review_http_error(error) from error


@router.post("/api/v1/relationship-reviews/{relationship_id}/revisions")
def create_relationship_revision(
    relationship_id: str,
    payload: RelationshipRevisionRequest,
    auditor: RelationshipReviewActor = Depends(relationship_auditor),
) -> dict[str, object]:
    try:
        return relationship_review_response(
            relationship_review_service().create_revision(
                relationship_id, payload.model_dump(), auditor
            )
        )
    except (KeyError, ValueError) as error:
        raise relationship_review_http_error(error) from error


# ── Export request / response models ───────────────────────────

class ExportRequest(BaseModel):
    model_config = ConfigDict(extra="forbid", strict=True)
    idempotency_key: str = Field(
        min_length=8, max_length=128, pattern=r"^[A-Za-z0-9][A-Za-z0-9_.-]*$"
    )


# ── Export endpoints ──────────────────────────────────────────

_shared_stores: dict[str, WorkbenchStore] = {}


def _get_shared_store(data_dir: str) -> WorkbenchStore:
    if data_dir not in _shared_stores:
        _shared_stores[data_dir] = WorkbenchStore(data_dir=Path(data_dir))
    return _shared_stores[data_dir]


@router.post("/api/v1/engagements/export", status_code=status.HTTP_201_CREATED)
def create_change_export(
    payload: ExportRequest,
    auditor: str = Depends(require_auditor),
) -> dict[str, object]:
    """Create a snapshot, detect changes, build and publish an export.

    Returns export metadata. The ZIP download is served via a separate
    GET endpoint.

    Idempotent — same idempotency_key returns the existing export.
    """
    from src.ace.workbench.engagement import EngagementService
    from src.ace.workbench.export_builder import generate_export_id
    from src.ace.workbench.notion_publisher import NotionPublisher

    data_dir = os.environ.get("ACE_DATA_DIR", "")
    store = _get_shared_store(data_dir) if data_dir else WorkbenchStore()

    # G0: reject real client
    try:
        current = EngagementService(store).current()
        if current is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No current engagement — cannot create export")
    except Exception as e:
        if "G0" in str(e):
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail=str(e))
        raise

    # Idempotency — check for existing export with this key
    with store.connect() as conn:
        existing = conn.execute(
            "SELECT DISTINCT export_id FROM change_records WHERE idempotency_key = ?",
            (payload.idempotency_key,),
        ).fetchone()
    if existing is not None:
        return {
            "export_id": existing["export_id"],
            "idempotent": True,
            "message": "Export already exists for this idempotency key",
        }

    # Create snapshot
    snapshot_result = store.create_snapshot(payload.idempotency_key)
    snapshot_id = str(snapshot_result["snapshot_id"])

    # Generate export ID with idempotency key for collision resistance
    export_id = generate_export_id(str(current.engagement_id), payload.idempotency_key)
    changes = store.detect_changes(snapshot_id, export_id, payload.idempotency_key)

    # Publish to fictional Notion
    publisher = NotionPublisher()
    publication_result = publisher.publish(
        idempotency_key=payload.idempotency_key,
        export_id=export_id,
    )

    return {
        "export_id": export_id,
        "snapshot_id": snapshot_id,
        "engagement_id": current.engagement_id,
        "change_count": len(changes),
        "warning_count": snapshot_result.get("warning_count", 0),
        "publication_id": publication_result.publication_id,
        "published": publication_result.published,
        "idempotent": False,
    }


@router.get("/api/v1/engagements/export/{export_id}")
def download_change_export(
    export_id: str,
    _: str = Depends(require_auditor),
) -> StreamingResponse:
    """Download the complete ACE change export ZIP for an export_id."""
    from io import BytesIO

    from src.ace.workbench.engagement import EngagementService
    from src.ace.workbench.export_builder import build_export_zip
    from src.ace.workbench.notion_publisher import NotionPublisher

    data_dir = os.environ.get("ACE_DATA_DIR", "")
    store = _get_shared_store(data_dir) if data_dir else WorkbenchStore()

    # Load change records for this export
    with store.connect() as conn:
        rows = conn.execute(
            "SELECT * FROM change_records WHERE export_id = ? ORDER BY timestamp",
            (export_id,),
        ).fetchall()
    if not rows:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Export not found")

    from src.ace.workbench.change_record import ChangeRecord

    changes = [
        ChangeRecord(
            change_id=row["change_id"],
            export_id=row["export_id"],
            record_id=row["record_id"],
            snapshot_id=row["snapshot_id"],
            evidence_id=row["evidence_id"],
            idempotency_key=row["idempotency_key"],
            timestamp=row["timestamp"],
            change_type=row["change_type"],
            record_type=row["record_type"],
            label=row["label"],
            detail=row["detail"],
        )
        for row in rows
    ]

    # Get snapshot info — use the engagement recorded in the snapshot,
    # not the current engagement (which may have changed since export).
    snapshot_id = rows[0]["snapshot_id"]
    with store.connect() as conn:
        snap_row = conn.execute(
            "SELECT * FROM snapshots WHERE snapshot_id = ?", (snapshot_id,)
        ).fetchone()
    if snap_row is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Snapshot not found")

    snapshot_engagement_id = snap_row["engagement_id"]

    # G0 check against the snapshot's engagement
    engagement = store.get_engagement(snapshot_engagement_id)
    if engagement.data_classification and engagement.data_classification not in ("FICTIONAL", "PUBLIC", "AUDITCO_OWNED"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="G0 blocks this export")

    engagement_title = engagement.title or "Untitled"

    warnings_list: list[dict[str, object]] = []
    try:
        snap_data = json.loads(snap_row["snapshot_data"])
        warnings_list = snap_data.get("warnings", [])
    except (json.JSONDecodeError, KeyError):
        pass

    # Build ZIP
    publisher = NotionPublisher()
    pub_result = publisher.publish(idempotency_key=rows[0]["idempotency_key"], export_id=export_id)

    zip_buf = build_export_zip(
        export_id=export_id,
        snapshot_id=snapshot_id,
        export_time=snap_row["created_at"] if snap_row else utc_now(),
        engagement_id=snapshot_engagement_id,
        engagement_title=engagement_title,
        changes=changes,
        warnings=warnings_list,
        excluded_count=0,
        publication_result=pub_result,
    )

    return StreamingResponse(
        BytesIO(zip_buf.getvalue()),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="ACE-change-export-{export_id}.zip"'
        },
    )


EVIDENCE_REVIEW_PAGE = """<!doctype html>
<html lang="en-AU"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Evidence Review | ACE Workbench</title><style>
:root { font-family: Arial, sans-serif; color: #17212b; background: #f4f7f8; } body { margin: 0; }
main { max-width: 960px; margin: auto; padding: 1rem; } section { background: #fff; border-radius: .5rem; box-shadow: 0 1px 3px #0002; margin: 1rem 0; padding: 1rem; }
form { display: grid; gap: .75rem; } label { display: grid; gap: .25rem; font-weight: bold; } input, select, textarea, button { font: inherit; min-height: 2.5rem; } textarea { min-height: 5rem; } button { background: #005a9c; color: #fff; border: 0; border-radius: .25rem; padding: .5rem; } .muted { color: #52606d; } .grid { display: grid; gap: .75rem; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); }
</style></head><body><main>
<p><a href="/workbench">Back To Field Capture Workbench</a></p><h1>Evidence Review</h1>
<p class="muted">G0 permits fictional, public, and AuditCo-owned Engagement data only.</p><p id="notice" role="status" aria-live="polite"></p>
<section><h2>Captured Evidence</h2><p id="evidence">Loading Evidence Item __EVIDENCE_ID__.</p><p id="engagement"></p><p><a id="media" hidden>Open Original Media</a></p><p id="unlinked"></p></section>
<section><h2>Review Suggestions</h2><p class="muted">Advisory only. ACE does not save, approve or apply suggestions.</p><p id="suggestions-status" role="status" aria-live="polite"></p><ul id="suggestions-list"></ul></section>
<section><h2>Source Context And Gaps</h2><form id="context-form"><div class="grid">
<label>Provider<input name="provider"></label><label>Origin<select name="origin"><option>RAW</option><option>DERIVED</option><option>AUDITOR_AUTHORED</option></select></label>
<label>Source Date<input name="source_date"></label><label>Source Version<input name="source_version"></label><label>Source Location<input name="source_location"></label><label>Freshness<select name="freshness"><option>CURRENT</option><option>STALE</option><option>SUPERSEDED</option><option>UNCERTAIN</option></select></label>
<label>Duplicate Evidence ID<input name="duplicate_evidence_id"></label><label>Source Evidence IDs<input name="source_evidence_ids" aria-describedby="sources-help"></label><label>Gap Status<select name="gap_status"><option>NOT_REQUESTED</option><option>REQUESTED_NOT_PROVIDED</option><option>UNAVAILABLE</option><option>STALE</option><option>INADEQUATE</option><option>NOT_APPLICABLE</option></select></label><label>Gap Materiality<select name="gap_materiality"><option value=""></option><option>MATERIAL</option><option>NOT_MATERIAL</option><option>UNDETERMINED</option></select></label></div>
<p id="sources-help" class="muted">Separate source Evidence IDs with commas.</p><label>Description<textarea name="description"></textarea></label><label>Limitations<textarea name="limitations"></textarea></label><label>Gap Explanation<textarea name="gap_explanation"></textarea></label><button id="context-save" disabled>Save Source Context</button></form></section>
<section><h2>Audit Questions</h2><p id="controls-notice" class="muted" hidden>No ACE Controls exist for this Engagement.</p><form id="question-form"><div class="grid"><label>Control ID<select name="control_id" id="control-id" required></select></label><label>Question Type<select name="question_type"><option>MAIN</option><option>IMPLEMENTATION</option><option>EFFECTIVENESS</option></select></label><label>MAIN Parent<select name="parent_question_id" id="parent-question"><option value=""></option></select></label></div><label>Question Text<textarea name="question_text" required></textarea></label><label>Purpose<textarea name="purpose" required></textarea></label><button id="question-submit" disabled>Create Audit Question</button></form>
<form id="version-form"><label>Question Version<select name="question_id" id="version-question"></select></label><label>Question Text<textarea name="question_text" required></textarea></label><label>Purpose<textarea name="purpose" required></textarea></label><button>Create Next Version</button></form></section>
<section><h2>Decisions And Proposed Links</h2><form id="decision-form"><div class="grid"><label>Exact Question Version<select name="question_version" id="decision-question"></select></label><label>Decision<select name="status"><option>APPROVED</option><option>REJECTED</option><option>CHANGES_REQUIRED</option></select></label></div><label>Reason<textarea name="reason" required></textarea></label><button>Record Decision</button></form>
<form id="link-form"><div class="grid"><label>Exact Question Version<select name="question_version" id="link-question"></select></label><label>Relevance<select name="relevance"><option>SUPPORTS</option><option>WEAKENS</option><option>CONTRADICTS</option></select></label></div><label>Reason<textarea name="reason" required></textarea></label><button>Propose Link</button></form><pre id="history"></pre></section>
<section><h2>Complete Review</h2><form id="complete-form"><label>Review Notes<textarea name="notes">Evidence review completed.</textarea></label><button>Complete Review</button></form><p id="completion"></p></section>
</main><script>
const evidenceId = '__EVIDENCE_ID__', notice = document.querySelector('#notice'); let state, decisionAttemptKey, completionAttemptKey;
const value = (form, name) => new FormData(form).get(name)?.trim() || null;
const request = async (url, method, body) => { const response = await fetch(url, {method, headers:{'Content-Type':'application/json'}, body: JSON.stringify(body)}); if (!response.ok) { const body = await response.json(); const detail = typeof body.detail === 'string' ? body.detail : JSON.stringify(body.detail); throw new Error(detail || 'The request failed.'); } return response.json(); };
const selectOptions = selector => { const select = document.querySelector(selector); select.innerHTML = state.question_versions.map(item => `<option value="${item.question_id}:${item.version}">${item.question_type} ${item.question_id} v${item.version}</option>`).join(''); };
async function load() { state = await request(`/workbench/api/v1/evidence/${evidenceId}/review`, 'GET'); document.querySelector('#evidence').textContent = `${state.evidence.evidence_id} — ${state.evidence.filename} (${state.evidence.status})`; document.querySelector('#engagement').textContent = `${state.engagement.title} — ${state.engagement.reference}`; const media = document.querySelector('#media'); media.hidden = !state.original_media_url; media.href = state.original_media_url || ''; document.querySelector('#unlinked').textContent = state.evidence.status === 'REVIEWED' && !state.proposed_links.length ? 'Reviewed Unlinked' : ''; document.querySelector('#completion').textContent = state.completion ? 'Review is complete.' : 'Review is pending.'; document.querySelector('#history').textContent = JSON.stringify({question_versions:state.question_versions, decisions:state.decisions, proposed_links:state.proposed_links}, null, 2); const parent = document.querySelector('#parent-question'); parent.innerHTML = '<option value=""></option>' + state.question_versions.filter(item => item.question_type === 'MAIN').map(item => `<option value="${item.question_id}">${item.question_id}</option>`).join(''); selectOptions('#version-question'); selectOptions('#decision-question'); selectOptions('#link-question'); }
const escapeText = value => String(value).replace(/[&<>"']/g, character => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[character]));
function hydrateSourceContext() { const form = document.querySelector('#context-form'), context = state.source_context; if (context) { for (const name of ['provider', 'origin', 'source_date', 'source_version', 'source_location', 'description', 'freshness', 'limitations', 'duplicate_evidence_id', 'gap_status', 'gap_explanation', 'gap_materiality']) form.elements[name].value = context[name] ?? ''; form.elements.source_evidence_ids.value = context.source_evidence_ids.join(', '); } document.querySelector('#context-save').disabled = false; }
function populateControls() { const select = document.querySelector('#control-id'), unavailable = !state.available_controls.length; select.innerHTML = state.available_controls.map(control => `<option value="${escapeText(control.control_id)}">${escapeText(control.control_id)} &mdash; ${escapeText(control.title)}</option>`).join(''); select.disabled = unavailable; document.querySelector('#question-submit').disabled = unavailable; document.querySelector('#controls-notice').hidden = !unavailable; }
const unhydratedLoad = load;
load = async () => { document.querySelector('#context-save').disabled = true; await unhydratedLoad(); populateControls(); hydrateSourceContext(); loadSuggestions(); };
function exact(selector) { const input = document.querySelector(selector), selected = value(input.form, input.name); if (!selected) throw new Error("Create an Audit Question version first."); const [question_id, question_version] = selected.split(':'); return {question_id, question_version:Number(question_version)}; }
function selectedQuestionId(form) { const selected = value(form, 'question_id'); if (!selected) throw new Error("Create an Audit Question version first."); return selected.split(':', 1)[0]; }
async function loadSuggestions() {
  try {
    const response = await fetch(`/workbench/api/v1/evidence/${evidenceId}/suggestions`);
    if (!response.ok) throw new Error('Suggestions unavailable.');
    const data = await response.json();
    const list = document.querySelector('#suggestions-list');
    const status = document.querySelector('#suggestions-status');
    list.innerHTML = '';
    if (!data.suggestions.length) { status.textContent = 'No review suggestions for this evidence item.'; return; }
    status.textContent = '';
    for (const s of data.suggestions) {
      const label = s.type === 'WARNING' ? '⚠️ WARNING' : '✅ FACT';
      const li = document.createElement('li');
      li.style.margin = '.75rem 0';
      li.innerHTML = `<strong>${escapeText(label)}</strong> &mdash; &ldquo;${escapeText(s.text)}&rdquo; <span class="muted">[${s.source_start}:${s.source_end}]</span>`;
      list.appendChild(li);
    }
  } catch (error) {
    document.querySelector('#suggestions-status').textContent = 'Suggestions are not available for this evidence item.';
    document.querySelector('#suggestions-list').innerHTML = '';
  }
}
document.querySelector('#context-form').onsubmit = async event => { event.preventDefault(); const form=event.currentTarget; try { await request(`/workbench/api/v1/evidence/${evidenceId}/review/context`, 'PUT', {...Object.fromEntries(new FormData(form)), source_evidence_ids:(value(form,'source_evidence_ids') || '').split(',').map(item=>item.trim()).filter(Boolean), gap_materiality:value(form,'gap_materiality')}); notice.textContent='Source context saved.'; await load(); } catch(error) { notice.textContent=error.message; } };
document.querySelector('#question-form').onsubmit = async event => { event.preventDefault(); const form=event.currentTarget; try { await request('/workbench/api/v1/audit-questions', 'POST', {...Object.fromEntries(new FormData(form)), engagement_id:state.engagement.engagement_id, parent_question_id:value(form,'parent_question_id')}); notice.textContent='Audit Question created.'; await load(); } catch(error) { notice.textContent=error.message; } };
document.querySelector('#version-form').onsubmit = async event => { event.preventDefault(); const form=event.currentTarget; try { const questionId=selectedQuestionId(form); await request(`/workbench/api/v1/audit-questions/${questionId}/versions`, 'POST', {question_text:value(form,'question_text'), purpose:value(form,'purpose')}); notice.textContent='Question version created.'; await load(); } catch(error) { notice.textContent=error.message; } };
document.querySelector('#decision-form').onsubmit = async event => { event.preventDefault(); const form=event.currentTarget; try { const item=exact('#decision-question'); decisionAttemptKey ??= `decision-${crypto.randomUUID()}`; await request(`/workbench/api/v1/audit-questions/${item.question_id}/decisions`, 'POST', {...item, decision_attempt_key:decisionAttemptKey, status:value(form,'status'), reason:value(form,'reason')}); decisionAttemptKey=null; notice.textContent='Decision recorded.'; await load(); } catch(error) { notice.textContent=error.message; } };
document.querySelector('#link-form').onsubmit = async event => { event.preventDefault(); const form=event.currentTarget; try { const item=exact('#link-question'); await request(`/workbench/api/v1/evidence/${evidenceId}/proposed-links`, 'POST', {...item, relevance:value(form,'relevance'), reason:value(form,'reason')}); notice.textContent='Link proposed.'; await load(); } catch(error) { notice.textContent=error.message; } };
document.querySelector('#complete-form').onsubmit = async event => { event.preventDefault(); const form=event.currentTarget; try { completionAttemptKey ??= `completion-${crypto.randomUUID()}`; await request(`/workbench/api/v1/evidence/${evidenceId}/review/complete`, 'POST', {completion_attempt_key:completionAttemptKey, notes:value(form,'notes')}); completionAttemptKey=null; notice.textContent='Review completed.'; await load(); } catch(error) { notice.textContent=error.message; } }; load().catch(error => notice.textContent=error.message);
</script></body></html>"""


WORKBENCH_PAGE = """<!doctype html>
<html lang="en-AU">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
  <meta name="theme-color" content="#005a9c">
  <title>ACE Field Capture Workbench</title>
  <style>
    :root { color-scheme: light; font-family: Arial, sans-serif; background: #f4f7f8; color: #17212b; }
    * { -webkit-tap-highlight-color: transparent; }
    body { margin: 0; overscroll-behavior-y: none; }
    main { max-width: 980px; min-height: 100vh; min-height: 100dvh; margin: auto; padding: max(1rem, env(safe-area-inset-top)) max(1rem, env(safe-area-inset-right)) max(calc(1rem + 56px + .75rem), calc(env(safe-area-inset-bottom) + 56px + .75rem)) max(1rem, env(safe-area-inset-left)); }
    h1 { margin-bottom: .25rem; } .grid { display: grid; gap: 1rem; grid-template-columns: repeat(auto-fit, minmax(130px, 1fr)); }
    .card { min-width: 0; background: #fff; border-radius: .5rem; padding: 1rem; box-shadow: 0 1px 3px #0002; }
    .capture-area { position: fixed; left: max(1rem, env(safe-area-inset-left)); right: max(1rem, env(safe-area-inset-right)); bottom: max(.75rem, env(safe-area-inset-bottom)); z-index: 1; max-width: 980px; margin: 0 auto; }
    .capture { display: flex; align-items: center; justify-content: center; box-sizing: border-box; min-height: 56px; background: #005a9c; color: white; font-size: 1.25rem; font-weight: bold; text-align: center; padding: 1rem; border-radius: .5rem; cursor: pointer; touch-action: manipulation; user-select: none; -webkit-user-select: none; }
    .capture.is-processing { background: #52606d; cursor: wait; }
    input[type=file] { display: none; } button, input, select, textarea { font-size: 16px; } button { min-height: 44px; background: #005a9c; color: white; border: 0; border-radius: .3rem; padding: .6rem; font: inherit; cursor: pointer; touch-action: manipulation; }
    li { margin: .5rem 0; } .muted { color: #52606d; } .notice { min-height: 1.5rem; } a { display: inline-flex; align-items: center; min-height: 44px; padding: 0 .25rem; color: #005a9c; touch-action: manipulation; }
    @media (max-width: 480px) { .grid { grid-template-columns: 1fr; } .card { padding: .75rem; } }
  </style>
</head>
<body>
  <main>
    <h1>Field Capture Workbench</h1>
    <p class="muted" id="engagement">__CURRENT_ENGAGEMENT__</p>
    <p><a href="/workbench/engagements/new">New Engagement</a></p>
    <div class="capture-area"><label class="capture" id="capture-control" aria-disabled="false">Capture Evidence<input id="capture" type="file" accept="image/*" capture="environment"></label><button id="retry-capture" type="button" hidden>Retry Capture</button><span class="muted" id="capture-engagement">__CURRENT_ENGAGEMENT__</span></div>
    <p class="notice" id="notice" aria-live="polite"></p>
    <section class="card"><h2>Mini Guide</h2><div class="grid" id="counts"></div></section>
    <section class="card"><h2>Fictional Chain</h2><p id="chain"></p></section>
    <section class="card"><h2>Recent Captures</h2><ul id="recent"></ul></section>
    <section class="card"><h2>Pending Review</h2><ul id="pending"></ul></section>
    <section class="card"><h2>Reviewed Unlinked</h2><ul id="reviewed-unlinked"></ul></section>
    <section class="card"><h2>Relationship Reviews</h2>__RELATIONSHIP_REVIEW_CONTENT__</section>
  </main>
  <script>
    const notice = document.querySelector('#notice');
    const captureInput = document.querySelector('#capture');
    const captureControl = document.querySelector('#capture-control');
    const retryCapture = document.querySelector('#retry-capture');
    let captureInProgress = false;
    let pendingCapture = null;
    const escapeText = (value) => String(value).replace(/[&<>"']/g, character => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[character]));
    async function loadWorkbench({ preserveNotice = false } = {}) {
      const response = await fetch('/workbench/summary');
      if (!response.ok) { if (!preserveNotice) notice.textContent = 'The workbench data is not available.'; return; }
      const data = await response.json();
      const current = data.current_engagement;
      const currentText = current ? `${current.title} — ${current.reference} — ${current.state}` : 'No READY_FOR_CAPTURE Engagement selected';
      document.querySelector('#engagement').textContent = currentText;
      document.querySelector('#capture-engagement').textContent = currentText;
      document.querySelector('#chain').textContent = data.chain.join(' → ');
      document.querySelector('#counts').innerHTML = Object.entries(data.counts).map(([name, value]) => `<div><strong>${value}</strong><br>${escapeText(name.replace('_', ' '))}</div>`).join('');
      document.querySelector('#recent').innerHTML = data.recent_captures.map(evidenceItem).join('');
      document.querySelector('#pending').innerHTML = data.pending_review.map(evidenceItem).join('');
      document.querySelector('#reviewed-unlinked').innerHTML = data.reviewed_unlinked.map(evidenceItem).join('');
    }
    function evidenceItem(item) {
      const media = item.media_url ? ` <a href="${item.media_url}">Preview</a>` : '';
      const action = item.status === 'PENDING_REVIEW' ? ` <button data-review="${item.evidence_id}">Review Evidence</button>` : ` <a href="/workbench/evidence/${encodeURIComponent(item.evidence_id)}/review">Review Evidence</a>`;
      return `<li><strong>${escapeText(item.evidence_id)}</strong> — ${escapeText(item.filename)} (${escapeText(item.status)})${media}${action}</li>`;
    }
    function readBase64(file) {
      return new Promise((resolve, reject) => {
        const reader = new FileReader();
        reader.onload = () => resolve(reader.result.split(',', 2)[1]);
        reader.onerror = () => reject(reader.error);
        reader.readAsDataURL(file);
      });
    }
    function setCaptureProcessing(isProcessing) {
      captureInput.disabled = isProcessing;
      retryCapture.disabled = isProcessing;
      captureControl.classList.toggle('is-processing', isProcessing);
      captureControl.setAttribute('aria-disabled', String(isProcessing));
      captureControl.setAttribute('aria-busy', String(isProcessing));
    }
    function captureAttemptKey() {
      return crypto.randomUUID ? crypto.randomUUID() : `capture-${Date.now()}-${Math.random().toString(16).slice(2)}`;
    }
    async function uploadCapture(payload) {
      notice.textContent = 'Uploading';
      const response = await fetch('/workbench/api/v1/evidence', {method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify(payload)});
      if (!response.ok) { pendingCapture = null; retryCapture.hidden = true; notice.textContent = 'Capture failed. The server did not accept the image.'; return; }
      const captured = (await response.json()).evidence;
      pendingCapture = null;
      retryCapture.hidden = true;
      notice.textContent = `Captured ${captured.evidence_id}.`;
    }
    captureInput.addEventListener('change', async event => {
      const file = event.target.files[0];
      if (!file || captureInProgress) { event.target.value = ''; return; }
      captureInProgress = true;
      setCaptureProcessing(true);
      let phase = 'read';
      try {
        notice.textContent = 'Preparing capture';
        const base64 = await readBase64(file);
        phase = 'upload';
        pendingCapture = {filename: file.name, media_type: file.type, data_base64: base64, capture_attempt_key: captureAttemptKey()};
        await uploadCapture(pendingCapture);
      } catch (error) {
        if (phase === 'read') { notice.textContent = 'Capture failed. The image file could not be read.'; }
        else { retryCapture.hidden = false; notice.textContent = 'Capture failed. The upload request could not be completed. Retry the upload.'; }
      } finally {
        captureInProgress = false;
        setCaptureProcessing(false);
        event.target.value = '';
        loadWorkbench({ preserveNotice: true });
      }
    });
    retryCapture.addEventListener('click', async () => {
      if (!pendingCapture || captureInProgress) return;
      captureInProgress = true;
      setCaptureProcessing(true);
      try { await uploadCapture(pendingCapture); }
      catch (error) { retryCapture.hidden = false; notice.textContent = 'Capture failed. The upload request could not be completed. Retry the upload.'; }
      finally { captureInProgress = false; setCaptureProcessing(false); loadWorkbench({ preserveNotice: true }); }
    });
    document.addEventListener('click', event => {
      const evidenceId = event.target.dataset.review; if (!evidenceId) return;
      window.location.assign(`/workbench/evidence/${evidenceId}/review`);
    });
    loadWorkbench();
  </script>
</body>
</html>"""


ENGAGEMENT_SETUP_PAGE = """<!doctype html>
<html lang="en-AU">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
  <meta name="theme-color" content="#005a9c">
  <title>New Engagement | ACE Workbench</title>
  <style>
    :root { font-family: Arial, sans-serif; color: #17212b; background: #f4f7f8; }
    body { margin: 0; }
    main { max-width: 780px; margin: auto; padding: max(1rem, env(safe-area-inset-top)) max(1rem, env(safe-area-inset-right)) max(1rem, env(safe-area-inset-bottom)) max(1rem, env(safe-area-inset-left)); }
    form, section { background: #fff; border-radius: .5rem; box-shadow: 0 1px 3px #0002; padding: 1rem; margin: 1rem 0; }
    .grid { display: grid; gap: 1rem; grid-template-columns: repeat(2, minmax(0, 1fr)); }
    label { display: grid; gap: .35rem; font-weight: bold; }
    input, textarea, select, button { box-sizing: border-box; font: inherit; font-size: 16px; min-height: 44px; }
    input, textarea, select { width: 100%; border: 1px solid #9aa5b1; border-radius: .25rem; padding: .5rem; }
    textarea { min-height: 100px; resize: vertical; }
    button { border: 0; border-radius: .3rem; background: #005a9c; color: white; padding: .6rem 1rem; cursor: pointer; touch-action: manipulation; }
    button[disabled] { background: #52606d; cursor: not-allowed; }
    .notice { min-height: 1.5rem; } .muted { color: #52606d; } .status { white-space: pre-wrap; }
    .offline-badge { display: inline-block; background: #f5a623; color: #fff; font-size: .75rem; font-weight: bold; padding: .2rem .5rem; border-radius: .25rem; margin-left: .5rem; vertical-align: middle; }
    @media (max-width: 560px) { .grid { grid-template-columns: 1fr; } }
  </style>
</head>
<body>
  <main>
    <p><a href="/workbench">Back To Field Capture Workbench</a></p>
    <h1>New Engagement <span id="offline-badge" class="offline-badge" hidden>OFFLINE_DRAFT</span></h1>
    <p class="muted">G0 permits fictional Engagements only. Do not enter real-client information.</p>
    <p class="muted" id="offline-status" hidden></p>
    <form id="engagement-form">
      <div class="grid">
        <label>Title<input name="title" maxlength="240"></label>
        <label>Engagement Reference<input name="reference" maxlength="120"></label>
        <label>Authority<input name="authority" maxlength="2000"></label>
        <label>Purpose<input name="purpose" maxlength="2000"></label>
        <label>Review Start Date<input name="review_start_date" type="date"></label>
        <label>Review End Date<input name="review_end_date" type="date"></label>
        <label>Evidence Cut-Off Date<input name="evidence_cut_off_date" type="date"></label>
        <label>Accountable Auditor<input name="accountable_auditor" maxlength="240"></label>
        <label>Data Classification<select name="data_classification"><option value="">Select classification</option><option value="FICTIONAL">FICTIONAL</option><option value="PUBLIC">PUBLIC</option><option value="AUDITCO_OWNED">AUDITCO_OWNED</option><option value="REAL_CLIENT">REAL_CLIENT</option></select></label>
        <label>Engagement Status<select name="is_fictional"><option value="">Select status</option><option value="true">Fictional</option><option value="false">Real-client</option></select></label>
      </div>
      <label>Scope<textarea name="scope"></textarea></label>
      <label>Exclusions<textarea name="exclusions"></textarea></label>
      <p><button type="submit">Save Draft</button></p>
    </form>
    <section aria-live="polite">
      <h2>Setup Summary</h2>
      <p class="notice" id="notice">Save a DRAFT for review.</p>
      <pre class="status" id="summary"></pre>
      <label><input id="confirm" type="checkbox"> I confirm the authority, scope, exclusions and fictional data boundary.</label>
      <p><button id="activate" type="button" disabled>Activate For Capture</button></p>
    </section>
  </main>
  <script>
    const form = document.querySelector('#engagement-form');
    const notice = document.querySelector('#notice');
    const summary = document.querySelector('#summary');
    const activate = document.querySelector('#activate');
    const confirm = document.querySelector('#confirm');
    const offlineBadge = document.querySelector('#offline-badge');
    const offlineStatus = document.querySelector('#offline-status');
    let engagementId = null;
    function creationAttemptKey() {
      return crypto.randomUUID ? crypto.randomUUID() : `attempt-${Date.now()}-${Math.random().toString(16).slice(2)}`;
    }
    function draftPayload() {
      const values = Object.fromEntries(new FormData(form).entries());
      return {
        ...values,
        creation_attempt_key: form.dataset.attemptKey || (form.dataset.attemptKey = creationAttemptKey()),
        data_classification: values.data_classification || null,
        is_fictional: values.is_fictional === '' ? null : values.is_fictional === 'true'
      };
    }
    function showRecord(record) {
      engagementId = record.engagement_id;
      summary.textContent = JSON.stringify(record, null, 2);
      activate.disabled = false;
    }
    // --- IndexedDB Offline Draft Storage ---
    const DB_NAME = 'ace-offline-drafts', STORE_NAME = 'engagement_drafts', DB_VERSION = 1;
    function openOfflineDB() {
      return new Promise((resolve, reject) => {
        if (!window.indexedDB) { reject(new Error('IndexedDB unavailable')); return; }
        const request = indexedDB.open(DB_NAME, DB_VERSION);
        request.onerror = function() { reject(request.error); };
        request.onsuccess = function() { resolve(request.result); };
        request.onupgradeneeded = function(event) {
          const db = event.target.result;
          if (!db.objectStoreNames.contains(STORE_NAME)) {
            db.createObjectStore(STORE_NAME, { keyPath: 'creation_attempt_key' });
          }
        };
      });
    }
    async function storeOfflineDraft(payload) {
      const db = await openOfflineDB();
      return new Promise((resolve, reject) => {
        const transaction = db.transaction(STORE_NAME, 'readwrite');
        transaction.onerror = function() { db.close(); reject(transaction.error); };
        transaction.oncomplete = function() { db.close(); };
        const store = transaction.objectStore(STORE_NAME);
        const record = { ...payload, stored_at: new Date().toISOString(), synced: false };
        const request = store.put(record);
        request.onsuccess = function() { resolve(record); };
        request.onerror = function() { reject(request.error); };
      });
    }
    async function getOfflineDrafts() {
      const db = await openOfflineDB();
      return new Promise((resolve, reject) => {
        const transaction = db.transaction(STORE_NAME, 'readonly');
        transaction.onerror = function() { db.close(); reject(transaction.error); };
        transaction.oncomplete = function() { db.close(); };
        const store = transaction.objectStore(STORE_NAME);
        const request = store.getAll();
        request.onsuccess = function() { resolve(request.result || []); };
        request.onerror = function() { reject(request.error); };
      });
    }
    async function removeOfflineDraft(creationAttemptKey) {
      const db = await openOfflineDB();
      return new Promise((resolve, reject) => {
        const transaction = db.transaction(STORE_NAME, 'readwrite');
        transaction.onerror = function() { db.close(); reject(transaction.error); };
        transaction.oncomplete = function() { db.close(); };
        const store = transaction.objectStore(STORE_NAME);
        const request = store.delete(creationAttemptKey);
        request.onsuccess = function() { resolve(); };
        request.onerror = function() { reject(request.error); };
      });
    }
    // --- G0 Client-Side Validation ---
    function validateOfflineG0(payload) {
      if (payload.data_classification === 'REAL_CLIENT' || payload.is_fictional === false) {
        throw new Error('G0 blocks real-client Engagement setup. Use FICTIONAL data only.');
      }
    }
    // --- Offline Detection And Sync ---
    function isOnline() { return navigator.onLine; }
    async function syncOfflineDrafts() {
      if (!isOnline()) return;
      const drafts = await getOfflineDrafts();
      if (!drafts.length) { offlineBadge.hidden = true; offlineStatus.hidden = true; return; }
      let syncedCount = 0, failed = 0;
      for (const draft of drafts) {
        try {
          const { stored_at, synced: _wasSynced, ...payload } = draft;
          const response = await fetch('/workbench/api/v1/engagements/sync', {
            method: 'POST', headers: {'Content-Type': 'application/json'},
            body: JSON.stringify(payload)
          });
          if (response.ok) {
            const data = await response.json();
            await removeOfflineDraft(draft.creation_attempt_key);
            if (syncedCount === 0 && !engagementId) showRecord(data.engagement);
            syncedCount++;
          } else {
            if (response.status === 409) {
              await removeOfflineDraft(draft.creation_attempt_key);
            }
            failed++;
          }
        } catch (error) { failed++; }
      }
      const remaining = await getOfflineDrafts();
      if (!remaining.length) {
        offlineBadge.hidden = true;
        offlineStatus.hidden = true;
        const parts = [`${syncedCount} offline draft${syncedCount !== 1 ? 's' : ''} synced.`];
        if (failed > 0) parts.push(`${failed} failed.`);
        notice.textContent = parts.join(' ');
      } else {
        offlineStatus.textContent = `${remaining.length} offline draft${remaining.length !== 1 ? 's' : ''} pending sync.`;
        offlineStatus.hidden = false;
        offlineBadge.hidden = false;
      }
    }
    function updateOfflineUI() {
      getOfflineDrafts().then(drafts => {
        const remaining = drafts.length;
        offlineBadge.hidden = remaining === 0;
        if (remaining > 0) {
          offlineStatus.textContent = `${remaining} offline draft${remaining !== 1 ? 's' : ''} pending sync.`;
          offlineStatus.hidden = false;
        } else { offlineStatus.hidden = true; }
      }).catch(function() {});
    }
    // Online detection: sync on reconnect
    window.addEventListener('online', function() {
      notice.textContent = 'Connection restored. Syncing offline drafts.';
      syncOfflineDrafts();
    });
    window.addEventListener('offline', function() {
      notice.textContent = 'Connection lost. Drafts will be saved offline.';
    });
    // Restore drafts on page load
    (async function restoreOfflineDrafts() {
      if (isOnline()) await syncOfflineDrafts();
      else updateOfflineUI();
    })();
    // --- Form Submission With Offline Support ---
    form.addEventListener('submit', async event => {
      event.preventDefault();
      activate.disabled = true;
      const payload = draftPayload();
      try { validateOfflineG0(payload); } catch (g0Error) {
        notice.textContent = g0Error.message;
        activate.disabled = false; return;
      }
      if (!isOnline()) {
        try {
          await storeOfflineDraft(payload);
          notice.textContent = 'Draft saved offline. It will sync when the connection returns.';
          updateOfflineUI();
        } catch (error) { notice.textContent = 'Offline draft could not be saved.'; }
        activate.disabled = false; return;
      }
      try {
        const response = await fetch('/workbench/api/v1/engagements', { method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify(payload) });
        if (!response.ok) { throw new Error('save failed'); }
        const data = await response.json();
        if (data.engagement.state === 'DRAFT') { confirm.checked = false; }
        showRecord(data.engagement);
        notice.textContent = 'The Engagement is saved as DRAFT. Review the setup before activation.';
      } catch (fetchError) {
        await storeOfflineDraft(payload);
        notice.textContent = 'Server unreachable. Draft saved offline. It will sync when the connection returns.';
        updateOfflineUI();
      }
      activate.disabled = false;
    });
    activate.addEventListener('click', async () => {
      if (!engagementId) return;
      const response = await fetch(`/workbench/api/v1/engagements/${engagementId}/activate`, { method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify({confirmed: confirm.checked}) });
      if (!response.ok) { notice.textContent = 'The Engagement was not activated. Complete the controlled setup and confirmation.'; return; }
      const data = await response.json();
      showRecord(data.engagement);
      notice.textContent = 'The Engagement is READY_FOR_CAPTURE and is the current capture context.';
    });
  </script>
</body>
</html>"""


ENGAGEMENT_SUMMARY_PAGE = """<!doctype html>
<html lang="en-AU"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Engagement Control Summary | ACE Workbench</title><style>
:root { font-family: Arial, sans-serif; color: #17212b; background: #f4f7f8; } body { margin: 0; }
main { max-width: 960px; margin: auto; padding: 1rem; } section { background: #fff; border-radius: .5rem; box-shadow: 0 1px 3px #0002; margin: 1rem 0; padding: 1rem; }
h2 { margin-top: 0; } table { border-collapse: collapse; width: 100%; } th, td { padding: .5rem; text-align: left; border-bottom: 1px solid #e1e4e8; } th { width: 14rem; color: #52606d; } .muted { color: #52606d; } ul { padding-left: 1.25rem; } ul li { margin: .25rem 0; }
</style></head><body><main>
<p><a href="/workbench">Back To Field Capture Workbench</a></p><h1>Engagement Control Summary</h1>
<p class="muted">Read-only view. G0 permits fictional, public, and AuditCo-owned Engagement data only.</p>
__CONTENT__
</main></body></html>"""
