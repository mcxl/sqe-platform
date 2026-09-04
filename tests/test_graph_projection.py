"""Focused tests for the read-only engagement graph projection (Phase 5)."""

import os
import re
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

from src.ace.app import app


def _credentials() -> tuple[str, str]:
    return ("auditor", "fictional-password")


@pytest.fixture
def client(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> TestClient:
    monkeypatch.setenv("ACE_AUDITOR_PASSWORD", "fictional-password")
    monkeypatch.setenv("ACE_DATA_DIR", str(tmp_path / "sqe-local-data"))
    return TestClient(app)


class TestGraphAuthentication:
    def test_unauthenticated_graph_returns_401(self, client: TestClient) -> None:
        assert client.get("/workbench/engagement/graph").status_code == 401

    def test_unauthenticated_export_returns_401(self, client: TestClient) -> None:
        assert client.get("/workbench/engagement/graph/export").status_code == 401


class TestGraphStructure:
    def test_graph_returns_html(self, client: TestClient) -> None:
        response = client.get("/workbench/engagement/graph", auth=_credentials())
        assert response.status_code == 200
        assert "text/html" in response.headers["content-type"]

    def test_graph_contains_engagement_id(self, client: TestClient) -> None:
        response = client.get("/workbench/engagement/graph", auth=_credentials())
        assert response.status_code == 200
        assert "ENG-FIC-0001" in response.text

    def test_graph_contains_nodes(self, client: TestClient) -> None:
        response = client.get("/workbench/engagement/graph", auth=_credentials())
        assert response.status_code == 200
        html = response.text
        # Check key node IDs appear
        for nid in ["EVD-FIC-0001", "OBL-FIC-0001", "RSK-FIC-0001", "CTL-FIC-0001",
                    "OWN-FIC-0001", "MATE-FIC-0001", "CON-FIC-0001"]:
            assert nid in html, f"Missing node {nid}"

    def test_graph_contains_svg(self, client: TestClient) -> None:
        response = client.get("/workbench/engagement/graph", auth=_credentials())
        assert response.status_code == 200
        assert "<svg" in response.text
        assert "</svg>" in response.text

    def test_graph_contains_records_table(self, client: TestClient) -> None:
        response = client.get("/workbench/engagement/graph", auth=_credentials())
        assert response.status_code == 200
        assert "<table>" in response.text
        assert "Type" in response.text
        assert "Label" in response.text

    def test_graph_contains_relationships_direction_table(self, client: TestClient) -> None:
        response = client.get("/workbench/engagement/graph", auth=_credentials())
        assert response.status_code == 200
        html = response.text
        assert "Relationships" in html, "Relationships section missing"
        assert "Direction" in html, "Direction header missing"
        # Verify each of the 6 approved edges renders with directional arrow
        assert "OBL-FIC-0001 → RSK-FIC-0001" in html
        assert "RSK-FIC-0001 → CTL-FIC-0001" in html
        assert "CTL-FIC-0001 → OWN-FIC-0001" in html
        assert "OWN-FIC-0001 → EVD-FIC-0001" in html
        assert "EVD-FIC-0001 → MATE-FIC-0001" in html
        assert "MATE-FIC-0001 → CON-FIC-0001" in html
        # CONTRA must NOT be in the relationships table (only approved edges)
        assert "CONTRA" not in html.split("Relationships")[1].split("Warnings")[0], (
            "CONTRA should not appear in the approved relationships table"
        )


class TestGraphEmptyState:
    def test_projection_empty_when_no_current_engagement(self, client: TestClient) -> None:
        """graph_projection returns None engagement_id when no engagement is active."""
        from src.ace.workbench.storage import WorkbenchStore

        data_dir = Path(os.environ["ACE_DATA_DIR"])
        store = WorkbenchStore(data_dir=data_dir)
        # Remove the current engagement to simulate no active engagement
        with store.connect() as conn:
            conn.execute("DELETE FROM current_engagement")
        result = store.graph_projection()
        assert result["engagement_id"] is None
        assert result["nodes"] == []
        assert result["edges"] == []
        assert result["warnings"] == []


class TestGraphConflicts:
    def test_conflict_relationship_in_warnings(self, client: TestClient) -> None:
        response = client.get("/workbench/engagement/graph", auth=_credentials())
        assert response.status_code == 200
        assert "CONTRA" in response.text, "CONTRA relationship must be visible"
        assert "Warnings" in response.text, "Warnings section must be present"


class TestGraphEvidenceNodes:
    def test_evidence_node_has_status(self, client: TestClient) -> None:
        response = client.get("/workbench/engagement/graph", auth=_credentials())
        assert response.status_code == 200
        assert "PENDING_REVIEW" in response.text

    def test_evidence_node_has_source_text_indicator(self, client: TestClient) -> None:
        # EVD-FIC-0001 has source text — the warning about "no source text" should NOT appear for it
        # (the INFO-level "no source text" warning only fires for evidence without source_text)
        response = client.get("/workbench/engagement/graph", auth=_credentials())
        assert response.status_code == 200
        # The evidence with source_text should not trigger the "no source text" warning
        assert "no source text" not in response.text.lower()


class TestGraphReadOnly:
    def test_graph_post_returns_405(self, client: TestClient) -> None:
        assert client.post("/workbench/engagement/graph", auth=_credentials()).status_code == 405

    def test_graph_put_returns_405(self, client: TestClient) -> None:
        assert client.put("/workbench/engagement/graph", auth=_credentials()).status_code == 405

    def test_graph_delete_returns_405(self, client: TestClient) -> None:
        assert client.delete("/workbench/engagement/graph", auth=_credentials()).status_code == 405

    def test_export_post_returns_405(self, client: TestClient) -> None:
        assert client.post("/workbench/engagement/graph/export", auth=_credentials()).status_code == 405


class TestXlsxExport:
    def test_export_returns_xlsx_content_type(self, client: TestClient) -> None:
        response = client.get("/workbench/engagement/graph/export", auth=_credentials())
        assert response.status_code == 200
        assert "spreadsheet" in response.headers.get("content-type", "")

    def test_export_has_attachment_header(self, client: TestClient) -> None:
        response = client.get("/workbench/engagement/graph/export", auth=_credentials())
        assert response.status_code == 200
        disposition = response.headers.get("content-disposition", "")
        assert "attachment" in disposition
        assert "ace-graph-ENG-FIC-0001" in disposition

    def test_export_sheets_exist(self, client: TestClient) -> None:
        from io import BytesIO

        from openpyxl import load_workbook

        response = client.get("/workbench/engagement/graph/export", auth=_credentials())
        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.content))
        sheet_names = wb.sheetnames
        assert "Records" in sheet_names
        assert "Evidence" in sheet_names
        assert "Relationships" in sheet_names
        assert "Warnings" in sheet_names
        assert "Read Me" in sheet_names

    def test_export_records_sheet_has_headers(self, client: TestClient) -> None:
        from io import BytesIO

        from openpyxl import load_workbook

        response = client.get("/workbench/engagement/graph/export", auth=_credentials())
        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.content))
        ws = wb["Records"]
        headers = [cell.value for cell in ws[1]]
        assert headers == ["ID", "Type", "Label", "Status"]

    def test_export_evidence_sheet_has_photo_metadata(self, client: TestClient) -> None:
        from io import BytesIO

        from openpyxl import load_workbook

        response = client.get("/workbench/engagement/graph/export", auth=_credentials())
        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.content))
        ws = wb["Evidence"]
        headers = [cell.value for cell in ws[1]]
        assert "Media Type" in headers
        assert "Photo Status" in headers

    def test_export_relationships_sheet_has_headers(self, client: TestClient) -> None:
        from io import BytesIO

        from openpyxl import load_workbook

        response = client.get("/workbench/engagement/graph/export", auth=_credentials())
        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.content))
        ws = wb["Relationships"]
        headers = [cell.value for cell in ws[1]]
        assert headers == ["Source ID", "Target ID", "Type", "Status"]

    def test_export_readme_has_export_id(self, client: TestClient) -> None:
        from io import BytesIO

        from openpyxl import load_workbook

        response = client.get("/workbench/engagement/graph/export", auth=_credentials())
        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.content))
        ws = wb["Read Me"]
        assert ws["A1"].value == "Export ID"
        assert "ENG-FIC-0001" in str(ws["B1"].value)

    def test_export_html_and_xlsx_use_same_engagement(self, client: TestClient) -> None:
        """Both graph HTML and XLSX come from the same graph_projection() call."""
        html_resp = client.get("/workbench/engagement/graph", auth=_credentials())
        xlsx_resp = client.get("/workbench/engagement/graph/export", auth=_credentials())
        assert html_resp.status_code == 200
        assert xlsx_resp.status_code == 200
        # Both mention the same engagement
        assert "ENG-FIC-0001" in html_resp.text
        assert "ENG-FIC-0001" in xlsx_resp.headers.get("content-disposition", "")

    def test_export_no_formulas(self, client: TestClient) -> None:
        from io import BytesIO

        from openpyxl import load_workbook

        response = client.get("/workbench/engagement/graph/export", auth=_credentials())
        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.content))
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    assert not (isinstance(cell.value, str) and cell.value.startswith("=")), \
                        f"Formula found in sheet {sheet_name}: {cell.value}"

    def test_export_no_photos_embedded(self, client: TestClient) -> None:
        from io import BytesIO

        from openpyxl import load_workbook

        response = client.get("/workbench/engagement/graph/export", auth=_credentials())
        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.content))
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            assert len(ws._images) == 0, f"Embedded image found in sheet {sheet_name}"


class TestApprovedRelationshipFiltering:
    """Only ACTIVE (non-CONTRA) relationships appear as graph edges."""

    def test_only_active_relationships_are_edges(self, client: TestClient) -> None:
        from src.ace.workbench.storage import WorkbenchStore

        store = WorkbenchStore(data_dir=Path(os.environ["ACE_DATA_DIR"]))
        result = store.graph_projection()
        edges = result["edges"]
        warnings = result["warnings"]
        # No CONTRA or OPEN edge in the edges list
        for edge in edges:
            assert edge["status"] == "ACTIVE", f"Edge with non-ACTIVE status: {edge}"
            assert edge["type"] != "CONTRA", f"CONTRA edge found: {edge}"
        # CONTRA relationship must be in warnings
        contra_warnings = [w for w in warnings if "CONTRA" in str(w.get("detail", ""))]
        assert len(contra_warnings) >= 1, "CONTRA relationship missing from warnings"

    def test_unapproved_in_warnings_not_edges(self, client: TestClient) -> None:
        from src.ace.workbench.storage import WorkbenchStore

        store = WorkbenchStore(data_dir=Path(os.environ["ACE_DATA_DIR"]))
        result = store.graph_projection()
        edges = result["edges"]
        warnings = result["warnings"]

        # Collect full (source, target, type, status) tuples from edges
        edge_keyed = {(str(e["source"]), str(e["target"]), str(e["type"]), str(e["status"])) for e in edges}

        # Collect unapproved relationships directly from the DB
        unapproved_keys: set[tuple[str, str, str, str]] = set()
        with store.connect() as conn:
            for row in conn.execute("SELECT * FROM relationships").fetchall():
                rtype = row["relationship_type"]
                rstatus = row["status"]
                if rstatus != "ACTIVE" or rtype == "CONTRA":
                    unapproved_keys.add((row["source_record_id"], row["target_record_id"], rtype, rstatus))

        assert unapproved_keys, "Expected at least one unapproved relationship in seed data"

        # No unapproved relationship (by full key) may appear in edges
        overlap = edge_keyed & unapproved_keys
        assert not overlap, (
            f"Unapproved relationships found in edges: {overlap}"
        )
        # All edges are approved
        assert all(e["status"] == "ACTIVE" for e in edges)
        assert all(e["type"] != "CONTRA" for e in edges)


class TestSnapshotIdentity:
    """HTML graph and XLSX export contain the same node and edge data."""

    def _get_snapshot(self, client: TestClient) -> tuple[list[dict], list[dict]]:
        """Return (nodes, edges) from a single graph_projection call."""
        from src.ace.workbench.storage import WorkbenchStore

        store = WorkbenchStore(data_dir=Path(os.environ["ACE_DATA_DIR"]))
        result = store.graph_projection()
        nodes: list[dict] = [{"id": n["id"], "type": n["type"]} for n in result["nodes"]]
        edges: list[dict] = [
            {"source": e["source"], "target": e["target"], "type": e["type"], "status": e["status"]}
            for e in result["edges"]
        ]
        return nodes, edges

    def test_snapshot_identity_hash(self, client: TestClient) -> None:
        import hashlib
        import json

        nodes, edges = self._get_snapshot(client)
        payload = json.dumps({"nodes": nodes, "edges": edges}, sort_keys=True)
        snapshot_hash = hashlib.sha256(payload.encode()).hexdigest()
        # Same call must produce identical hash
        nodes2, edges2 = self._get_snapshot(client)
        payload2 = json.dumps({"nodes": nodes2, "edges": edges2}, sort_keys=True)
        snapshot_hash2 = hashlib.sha256(payload2.encode()).hexdigest()
        assert snapshot_hash == snapshot_hash2, "Snapshot hash must be deterministic"

    def test_html_and_xlsx_share_same_node_ids(self, client: TestClient) -> None:
        from io import BytesIO

        from openpyxl import load_workbook

        nodes, _ = self._get_snapshot(client)
        xlsx_resp = client.get("/workbench/engagement/graph/export", auth=_credentials())
        assert xlsx_resp.status_code == 200

        wb = load_workbook(BytesIO(xlsx_resp.content))
        ws = wb["Records"]
        xlsx_ids = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                xlsx_ids.add(str(row[0]))
        projection_ids = {str(n["id"]) for n in nodes}

        missing = projection_ids - xlsx_ids
        extra = xlsx_ids - projection_ids
        assert not missing, f"Node IDs in projection but not XLSX: {missing}"
        assert not extra, f"Node IDs in XLSX but not projection: {extra}"
        assert projection_ids == xlsx_ids

    def test_html_and_xlsx_share_same_edge_pairs(self, client: TestClient) -> None:
        from io import BytesIO

        from openpyxl import load_workbook

        _, edges = self._get_snapshot(client)
        xlsx_resp = client.get("/workbench/engagement/graph/export", auth=_credentials())
        assert xlsx_resp.status_code == 200

        wb = load_workbook(BytesIO(xlsx_resp.content))
        ws = wb["Relationships"]
        xlsx_pairs: set[tuple[str, str]] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                xlsx_pairs.add((str(row[0]), str(row[1])))
        projection_pairs = {(str(e["source"]), str(e["target"])) for e in edges}

        missing = projection_pairs - xlsx_pairs
        extra = xlsx_pairs - projection_pairs
        assert not missing, f"Edge pairs in projection but not XLSX: {missing}"
        assert not extra, f"Edge pairs in XLSX but not projection: {extra}"
        assert projection_pairs == xlsx_pairs


class TestRelationshipDirection:
    """Every edge has explicit direction with valid source and target nodes."""

    def test_every_edge_source_and_target_exist(self, client: TestClient) -> None:
        from src.ace.workbench.storage import WorkbenchStore

        store = WorkbenchStore(data_dir=Path(os.environ["ACE_DATA_DIR"]))
        result = store.graph_projection()
        node_ids = {str(n["id"]) for n in result["nodes"]}

        for edge in result["edges"]:
            assert str(edge["source"]) in node_ids, (
                f"Edge source {edge['source']} not in nodes"
            )
            assert str(edge["target"]) in node_ids, (
                f"Edge target {edge['target']} not in nodes"
            )
            assert str(edge["source"]) != str(edge["target"]), (
                f"Self-referencing edge: {edge['source']} → {edge['target']}"
            )

    def test_every_edge_has_explicit_direction(self, client: TestClient) -> None:
        from src.ace.workbench.storage import WorkbenchStore

        store = WorkbenchStore(data_dir=Path(os.environ["ACE_DATA_DIR"]))
        result = store.graph_projection()

        for edge in result["edges"]:
            assert "source" in edge, f"Edge missing 'source': {edge}"
            assert "target" in edge, f"Edge missing 'target': {edge}"
            assert isinstance(edge["source"], str), f"Source not a string: {edge['source']}"
            assert isinstance(edge["target"], str), f"Target not a string: {edge['target']}"
            assert edge["source"] != "", "Edge has empty source"
            assert edge["target"] != "", "Edge has empty target"

    def test_all_active_edges_have_direction(self, client: TestClient) -> None:
        from src.ace.workbench.storage import WorkbenchStore

        store = WorkbenchStore(data_dir=Path(os.environ["ACE_DATA_DIR"]))
        result = store.graph_projection()
        edges = result["edges"]

        # Seed data has 7 relationships, 6 ACTIVE non-CONTRA → expect 6 edges
        assert len(edges) == 6, f"Expected 6 approved edges, got {len(edges)}"
        # Every edge is directional
        for edge in edges:
            assert edge["source"] != edge["target"]
